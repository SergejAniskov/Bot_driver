from datetime import datetime, timedelta, date
import asyncio
import logging
import sys
import keyboards.keyboard
import sqlite
import emoji
import calendar

import openpyxl
from openpyxl.worksheet import worksheet
from datetime import datetime
from aiogram.exceptions import TelegramBadRequest
from aiogram import types
from os import getenv
from re import Match
from magic_filter import RegexpMode
from aiogram import Bot, Dispatcher, F, Router, html
from aiogram.enums import ParseMode
from aiogram.filters import Command, CommandStart
from aiogram.fsm.context import FSMContext
from contextlib import suppress
from aiogram.fsm.state import State, StatesGroup

from aiogram.types import (
    KeyboardButton,
    Message,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
)




TOKEN = getenv("BOT_TOKEN")

form_router = Router()


class Form(StatesGroup):
    name = State()
    like_bots = State()
    language = State()
    cash = State()
    id_us = State()
    cat_new = State()
    mes_id = State()
    coments = State()
    fuel_vol = State()

class Dates(StatesGroup):
    day = State()
    month = State()
    yer = State()
    itog = State()

class Fuels(StatesGroup):
    fuel_vol = State()
class Flight(StatesGroup):
    name = State()
    status = State()
    cars = State()
    mileage = State()
    fuel = State()
class Flight_exit(StatesGroup):
    status = State()
    name = State()
    finish_odometr = State()
    fuel_residues = State()

class orders_gen(StatesGroup):
    namber = State()



now = datetime.now().strftime("%Y-%m-%d %H:%M")
now2 = datetime.now().strftime("%d.%m.%Y")
now1 = datetime.now().strftime("%Y-%m-%d %H:%M:%S")


# categories - Управление категориями
# flight - Управление Рейсами
# cancel - Выход


async def on_startup(_):
    await sqlite.db_connect()
    print("Подключение БД ОК")


def transform_date(date):
    months = ['января', 'февраля', 'марта', 'апреля', 'мая', 'июня',
              'июля', 'августа', 'сентября', 'октября', 'ноября', 'декабря']
    day, month, year = now2.split('.')
    return f'{day} {months[int(month) - 1]} {year} года'

def transform_date_new(date):
    months = ['января', 'февраля', 'марта', 'апреля', 'мая', 'июня',
              'июля', 'августа', 'сентября', 'октября', 'ноября', 'декабря']
    year, month, day = date.split('-')
    return f'{day} {months[int(month) - 1]} {year} года'



############ Нажатие кнопки Создать Эксель
@form_router.callback_query(keyboards.keyboard.Flight_Callback.filter(F.action == "order_flights_exel"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.Flight_Callback
) -> None:
    flight_id = callback_data.id_fl


    await exel_creat(flight_id)
    await callback.message.edit_text("Создан отчёт")
    await callback.answer()



# # Обработка доставок
@form_router.message(Command("delivery"))
async def buy(message: types.Message, state: FSMContext):
    await message.answer(
        f"Управление доставками",
        reply_markup=keyboards.keyboard.get_dev_pan()
    )


# # Обработка Эксель
@form_router.message(Command("exel"))
async def buy(message: types.Message, state: FSMContext):
    await state.set_state(orders_gen.namber)

    await message.answer("Введи номер рейса")

@form_router.message(orders_gen.namber)
async def get_info_m(message: types.Message, state: FSMContext) -> None:
    id = message.text
    await exel_creat(id)
    await state.clear()
    await message.answer("Ok")




# Обработка Эксель
async def exel_creat(id_fl: int):

    filename = "templete.xlsx"
    book = openpyxl.load_workbook(filename=filename)
    sheet: worksheet = book.worksheets[0]

    incr_db = await sqlite.db_sel_other_shapka(id_fl)
    auto = ""
    name = ""
    for i in incr_db:
        sheet["B1"].value = i[0]
        name = i[1]
        auto = i[1]

        sheet["B3"].value = auto +" "+ i[2]
        sheet["B4"].value = i[3]

        sheet["J2"].value = i[4]
        date_in = datetime.strptime((i[4]), '%Y-%m-%d %H:%M')
        date_ex = datetime.strptime((i[5]), '%Y-%m-%d %H:%M')
        sheet["L2"].value = i[5]
        s = i[5]

        # print(s[:-6])
        name = name + "От" + s[:-6]
        sheet["J3"].value = i[6]
        sheet["L3"].value = i[7]
        sheet["J4"].value = i[8]
        sheet["L4"].value = i[9]
        sheet["E4"].value = i[10]
        delta_days = (date_ex - date_in).days
        sheet["N2"].value = delta_days



    incr_db = await sqlite.db_sel_tr_pay(id_fl,"incr")
    fuel_db = await sqlite.db_sel_tr_fuels(id_fl)
    other_db = await sqlite.db_sel_tr_other(id_fl)



# # Заполняем поступление
#     print(incr_db)
    nomer = 7
    for i in incr_db:
        pay = i[1]

        if 'карт' in pay:
            sheet.cell(row=nomer, column=2).value = i[2]
        else:
            sheet.cell(row=nomer, column=3).value = i[2]

        date = datetime.strptime((i[3]), '%Y-%m-%d %H:%M:%S')
        sheet.cell(row=nomer, column=1).value = date.strftime('%d-%m-%y %H:%M')
        nomer = nomer + 1


# Заполняем топливо
    nomer = 7
    for i in fuel_db:
        date = datetime.strptime((i[1]), '%Y-%m-%d %H:%M:%S')
        sheet.cell(row=nomer, column=6).value = date.strftime('%d-%m-%y %H:%M')
        sheet.cell(row=nomer, column=7).value = i[2]
        sheet.cell(row=nomer, column=8).value = i[3]
        nomer = nomer + 1

# Заполняем прочие расходы
    nomer = 7
    coll = 0
    categ = ""



    for i in other_db:

        if nomer > 24:
            nomer = 7
            coll = 3

        date = datetime.strptime((i[1]), '%Y-%m-%d %H:%M:%S')
        sheet.cell(row=nomer, column=9+coll).value = value = date.strftime('%d-%m-%y %H:%M')
        categ = emoji.replace_emoji(i[2], replace='')
        sheet.cell(row=nomer, column=11+coll).value = i[3]

        if i[4] != None:
            categ = categ + " " + f'"{emoji.replace_emoji(i[4], replace="")}"'
            sheet.cell(row=nomer, column=10 + coll).value = categ
        else:
            sheet.cell(row=nomer, column=10 + coll).value = categ
        nomer = nomer + 1




    # book.save(f'{name}.xlsx')

    filepath = f'Отчеты/{name}.xlsx'
    book.save(filepath)


    # await message.answer(f"Вставил")


def date_diff(date1, date2):
    d1 = date.strptime(date1, '%Y-%m-%d')
    t1 = datetime.strptime(date1 + ' ' + date1.split()[-1], '%Y-%m-%d %H:%M').time()
    d2 = date.strptime(date2, '%Y-%m-%d')
    t2 = datetime.strptime(date2 + ' ' + date2.split()[-1], '%Y-%m-%d %H:%M').time()

    delta = d2 - d1 - timedelta(hours = t2.hour-t1.hour , minutes = t2.minute-t1.minute )
    return delta.days


# Обработка выбора меню категорииЯ
@form_router.callback_query(keyboards.keyboard.NumbersCallbackFactory.filter(F.action == "select"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.NumbersCallbackFactory,
        state: FSMContext
    ) -> None:

    await state.update_data(name=callback_data.val)
    data = await state.update_data(language=callback_data.value)
    await state.update_data(mes_id=callback.message.message_id)

    cash = data["cash"]
    sell = data["language"]
    cat = data["name"]
    us_id = data["id_us"]
    flight = data["id_flight"]


    if sell == 123:


        await callback.message.edit_text(
            "💬 Введите, количество литров."
        )
        await state.set_state(Form.fuel_vol)
        await state.update_data(mes_text=callback.message)


    else:
        await select_cat(callback.message, cat, sell, cash, us_id, 0, flight,"")
        await state.clear()
        print("mes_ID")
        print(callback.message.message_id)

    await callback.answer()


@form_router.message(Form.fuel_vol)
async def get_info_m(message: types.Message, state: FSMContext) -> None:
    litr = message.text
    litr = litr.replace(",", ".")
    try:
        float(litr)
        data = await state.get_data()
        cash = data["cash"]
        sell = data["language"]
        cat = data["name"]
        us_id = data["id_us"]
        flight = data["id_flight"]
        mes_text = data["mes_text"]

        await select_cat(mes_text, cat, sell, cash, us_id, 0, flight, litr)
        await state.clear()

    except ValueError:
        await state.set_state(Form.fuel_vol)
        await message.reply(
            f'Введённое значение не является числом!\n💬 Введите, количество литров. в формате "112.5"'
        )



# Обработка del trancaction
@form_router.callback_query(keyboards.keyboard.Itog_Callback.filter(F.action == "tr_cancel"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.Itog_Callback
) -> None:

    id_tranz = callback_data.id_tranz
    user = callback_data.val
    data = await sqlite.sel_tranc(id_tranz,user)

    cash = ""
    cat = ""
    date = ""
    for i in data:
        cash = i[0]
        cat = i[1]
        date = i[2]

    await sqlite.db_del_transactions(id_tranz, user)
    await callback.message.edit_text(
        f"Добавлено <b>{cat}</b> ₽ в категорию <b>{cash}</b> операция расход \n {date} \n---\n🚫 Отменено"
    )



# Обработка редактирование даты
@form_router.callback_query(keyboards.keyboard.Itog_Callback.filter(F.action == "tr_date"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.Itog_Callback,
        state: FSMContext
) -> None:
     # mes_id = mes_id,
    id_tranz = callback_data.id_tranz
    us_id = callback_data.val
    # mes_id = callback.data.mes_id

    data = await sqlite.sel_tranc(id_tranz,us_id)

    cash = ""
    cat = ""
    date = ""
    for i in data:
        cash = i[0]
        cat = i[1]
        date = i[2]

    dates = datetime.strptime(date, '%Y-%m-%d %H:%M:%S')
    ss = dates.strftime('%d-%m-%y %H:%M')
    month = dates.month
    year = dates.year
    days_in_month = calendar.monthrange(dates.year, dates.month)[1]  # 28
    text_mes = f"Добавлено <b>{cat}</b> ₽ в категорию <b>{cash}</b> операция расход \n {date} "
    await state.update_data(text_mes=text_mes)
    await state.update_data(dates=date)
    await state.update_data(dates_in=date)
    await state.update_data(id_tranz=id_tranz)
    await state.update_data(us_id=us_id)

    await callback.message.edit_text(
        text_mes + "\n<b>Изменить дату на:</b>",
        reply_markup=keyboards.keyboard.get_key_date(id_tranz,dates.day,month,year)
    )


# Обработка нопки назад к дате
@form_router.callback_query(keyboards.keyboard.Itog_Callback.filter(F.action == "pr_date"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.Itog_Callback,
        state: FSMContext
) -> None:

    data = await state.get_data()
    date = data["dates"]
    id_tranz = data["id_tranz"]
    text_mes = data["text_mes"]
    dates = datetime.strptime(date, '%Y-%m-%d %H:%M:%S')


    await callback.message.edit_text(
         text_mes + "\n<b>Изменить дату на:</b>",
         reply_markup=keyboards.keyboard.get_key_date(id_tranz, dates.day, dates.month, dates.year)
    )

###################################################
# Обработка кнопки готово дата
@form_router.callback_query(keyboards.keyboard.Itog_Callback.filter(F.action == "dates_finish"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        state: FSMContext
) -> None:

    data = await state.get_data()
    text_mes = data["text_mes"]
    id_tranz = data["id_tranz"]
    date = data["dates"]
    date_in = data["dates_in"]
    us_id = data["us_id"]

    dates = datetime.strptime(date, '%Y-%m-%d %H:%M:%S')
    dates_in = datetime.strptime(date_in, '%Y-%m-%d %H:%M:%S')

    if dates == dates_in:

        text_mes = text_mes[:-20] + f" {transform_date_new(date[:-9])} {date[11:]}"
    else:
        await sqlite.update_trans_date(date,id_tranz)
        text_mes = text_mes[:-20] + f"Дата изменена на {transform_date_new(date[:-9])} {date_in[11:]}"

    await callback.message.edit_text(
        text_mes,
        reply_markup=keyboards.keyboard.get_keyboard(id_tranz, callback.message.message_id, us_id)
    )
    await state.clear()


###################################################
# Обработка редактирование года
@form_router.callback_query(keyboards.keyboard.Itog_Callback.filter(F.action == "dates_year"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        state: FSMContext
) -> None:

    data = await state.get_data()
    text_mes = data["text_mes"]
    id_tranz = data["id_tranz"]
    date = data["dates"]
    us_id = data["us_id"]

    dates = datetime.strptime(date, '%Y-%m-%d %H:%M:%S')

    await callback.message.edit_text(
        text_mes + "\n<b>Выбери год:</b>",
        reply_markup=keyboards.keyboard.get_key_year(dates.year,id_tranz,us_id)
    )

# Установка выбраного года
@form_router.callback_query(keyboards.keyboard.Itog_Callback.filter(F.action == "year_set"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.Itog_Callback,
        state: FSMContext
) -> None:

    data = await state.get_data()
    date = data["dates"]
    id_tranz = data["id_tranz"]
    text_mes = data["text_mes"]
    dates = datetime.strptime(date, '%Y-%m-%d %H:%M:%S')
    select_year = callback_data.val
    d_new = dates.replace(year=select_year)
    await state.update_data(dates=d_new.strftime('%Y-%m-%d %H:%M:%S'))

    await callback.message.edit_text(
        text_mes+ "\n<b>Изменить дату на:</b>",
        reply_markup=keyboards.keyboard.get_key_date(id_tranz,d_new.day,d_new.month,d_new.year)
    )




###################################################
# Обработка редактирование месяца
@form_router.callback_query(keyboards.keyboard.Itog_Callback.filter(F.action == "dates_month"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        state: FSMContext
) -> None:

    data = await state.get_data()
    text_mes = data["text_mes"]
    id_tranz = data["id_tranz"]
    us_id = data["us_id"]

    await callback.message.edit_text(
        text_mes + "\n<b>Выбери месяц:</b>",
        reply_markup=keyboards.keyboard.get_key_month(id_tranz,us_id)
    )

# Установка выбраного месяца
@form_router.callback_query(keyboards.keyboard.Itog_Callback.filter(F.action == "month_set"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.Itog_Callback,
        state: FSMContext
) -> None:

    data = await state.get_data()
    date = data["dates"]
    id_tranz = data["id_tranz"]
    text_mes = data["text_mes"]
    dates = datetime.strptime(date, '%Y-%m-%d %H:%M:%S')
    select_month = callback_data.val
    d_new = dates.replace(month=select_month)
    await state.update_data(dates=d_new.strftime('%Y-%m-%d %H:%M:%S'))

    await callback.message.edit_text(
        text_mes + "\n<b>Изменить дату на:</b>",
        reply_markup=keyboards.keyboard.get_key_date(id_tranz,d_new.day,d_new.month,d_new.year)
    )


# Обработка редактирование даты
@form_router.callback_query(keyboards.keyboard.Itog_Callback.filter(F.action == "dates_days"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        state: FSMContext
) -> None:

    data = await state.get_data()
    text_mes = data["text_mes"]
    date = data["dates"]
    id_tranz = data["id_tranz"]
    us_id = data["us_id"]
    dates = datetime.strptime(date, '%Y-%m-%d %H:%M:%S')
    days_in_month = calendar.monthrange(dates.year, dates.month)[1]  # 28
    await callback.message.edit_text(
        text_mes + "\n<b>Выбери день:</b>",
        reply_markup=keyboards.keyboard.get_key_days(days_in_month,id_tranz,us_id)
    )

# Установка выбраного дня
@form_router.callback_query(keyboards.keyboard.Itog_Callback.filter(F.action == "days_set"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.Itog_Callback,
        state: FSMContext
) -> None:

    data = await state.get_data()
    date = data["dates"]
    id_tranz = data["id_tranz"]
    text_mes = data["text_mes"]
    dates = datetime.strptime(date, '%Y-%m-%d %H:%M:%S')
    select_day = callback_data.val
    d_new = dates.replace(day=select_day)
    await state.update_data(dates=d_new.strftime('%Y-%m-%d %H:%M:%S'))

    await callback.message.edit_text(
        text_mes + "\n<b>Изменить дату на:</b>",
        reply_markup=keyboards.keyboard.get_key_date(id_tranz,d_new.day,d_new.month,d_new.year)
    )




# Обработка Коментарий trancaction
@form_router.callback_query(keyboards.keyboard.Itog_Callback.filter(F.action == "tr_caments"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.Itog_Callback,
        state: FSMContext
) -> None:

    id_tranz = callback_data.id_tranz
    user = callback_data.val
    data = await sqlite.sel_tranc(id_tranz,user)

    cash = ""
    cat = ""
    date = ""
    for i in data:
        cash = i[0]
        cat = i[1]
        date = i[2]

    await callback.message.edit_text(
        f"💬 Введите, пожалуйста, комментарий."
    )
    await state.set_state(Form.coments)

    await state.update_data(cat=cat)
    await state.update_data(us_id=user)
    await state.update_data(cash=cash)
    await state.update_data(dates=date)
    await state.update_data(id_tr=id_tranz)


##################################################
############ Коментарий

@form_router.message(Form.coments)
# async def process_name(callback: types.CallbackQuery, state: FSMContext) -> None:
async def process_name(message: Message, state: FSMContext) -> None:
    coments = message.text

    data = await state.get_data()
    cash = data["cash"]
    cat = data["cat"]
    us_id = data["us_id"]
    dates = data["dates"]

    idi = data["id_tr"]



    await sqlite.update_trans_com(coments, idi)

    await message.answer( text=f"👌 Комментарий успешно добавлен.\n Добавлено {cash} ₽ в категорию {cat} \n  операция расход \n {dates}, <i><b>{coments}</b></i>",
          reply_markup=keyboards.keyboard.get_keyboard(idi, message.message_id, us_id)
    )
    await state.clear()
    print(f"ID Mess  coments {coments}")




# Обработка del trancaction
@form_router.callback_query(keyboards.keyboard.Itog_Callback.filter(F.action == "inc_tr_kl"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.Itog_Callback,
        state: FSMContext
) -> None:
    data = await state.get_data()
    cash = data["cash"]
    sell = data["language"]
    cat = data["name"]
    us_id = data["id_us"]
    flight = data["id_flight"]
    name_flight = data["name_flight"]

    znac = callback_data.val
    mes_id = data["mes_id"]

    # print(f"Del_tranc {znac}")
    if znac == 1:
        print(f"date edit {data}")




async def select_cat(message: types.Message, cat: str, sellect: str, cash: str, us_id: int, edit: int, flight: int, fuel: int):
    with suppress(TelegramBadRequest):

        if edit == 0:
            dann = await sqlite.db_insert_transactions(cat, us_id, cash, now1, flight, fuel)
            idi = 0
            users = 0
            for i in dann:
                idi = i[0]

            await message.edit_text(
                f"Добавлено {cash} ₽ в категорию {cat} операция расход \n {transform_date(now1)} ",
                reply_markup=keyboards.keyboard.get_keyboard(idi, message.message_id, us_id)
            )
            await message.edit_text(
                f"Добавлено {cash} ₽ в категорию {cat} операция расход \n {transform_date(now1)} ",
                reply_markup=keyboards.keyboard.get_keyboard(idi,message.message_id,us_id)
            )
            # print(f"id {id}")
            # print(f"message_id {message.message_id}")




@form_router.message(CommandStart())
async def command_start(message: Message, state: FSMContext) -> None:
    id = message.from_user.id
    data = await sqlite.db_sel_driver(id)
    user = ""
    if len(data) != 0:
        for i in data:
            user = i[0]
        await message.answer(f"Приветствую <b>{user}</b>!\nВы зарегистрированы в системе, можете продолжать работать.")
    else:
        await state.set_state(Form.name)
        await message.answer(
            "Добро пожаловать!\nПредставьтесь полным ФИО, будем использовать для отчёта.\n"
            "   Этот бот может помочь водителю в учете доходов и расходов, а также создании отчетов о рейсах.",
            reply_markup=ReplyKeyboardRemove(),
        )

@form_router.message(Form.name)
async def handle_code(message: types.Message, state: FSMContext) -> None:
    user_name = message.text
    user_id = message.from_user.id
    await sqlite.db_insert_driver(user_id,user_name)
    await state.clear()
    await message.answer(
        f"Поздравляю <b>{user_name}</b>!\nВы зарегистрированы в системе, можете продолжать работать.",
        reply_markup=ReplyKeyboardRemove(),
    )
    # await message.reply("Я тебя не понимаю :(")


# ОБРАБОТКА К Приходу
@form_router.callback_query(keyboards.keyboard.NumbersCallbackFactory.filter(F.action == "incr"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.NumbersCallbackFactory,
        state: FSMContext
) -> None:
    data = await state.get_data()
    cash = data["cash"]
    name_flight = data["name_flight"]
    users_id = callback.message.chat.id
# Запрос категории

    data = await sqlite.db_sel_cat(users_id, "incr")

    await callback.message.edit_text(
        f"{name_flight}\nВыбери куда записать: {cash} ₽",
        reply_markup=keyboards.keyboard.gen_markup_deb(data)
    )

# ОБРАБОТКА К Расходу
@form_router.callback_query(keyboards.keyboard.NumbersCallbackFactory.filter(F.action == "decr"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.NumbersCallbackFactory,
        state: FSMContext
) -> None:
    data = await state.get_data()
    cash = data["cash"]
    users_id = callback.message.chat.id

# Запрос категории
    data= await sqlite.db_sel_cat(users_id, "decr")

    await callback.message.edit_text(
        f"Выбери куда записать: {cash} ₽",
        reply_markup=keyboards.keyboard.gen_markup(data)
    )

# Запрос списка категорий для редактирование
@form_router.message(Command("categories"))
async def buy(message: types.Message):

    data= await sqlite.db_sel_edit_cat(message.chat.id)

    if len(data) != 0:
        await message.answer(f"Список категорий", reply_markup=keyboards.keyboard.gen_list_cat(data))
    else:
        await message.answer(
            f"В вашем списке категорий не обнаружено, предлагаю создать свои или заполнить их по умолчанию.",
            reply_markup=keyboards.keyboard.gen_list_cat_nul()
        )
# Заполняем категории по умолчанию
@form_router.callback_query(keyboards.keyboard.NumbersCallbackFactory.filter(F.action == "ins_cat_def"))
async def command_start(callback: types.CallbackQuery) -> None:
    dan = await sqlite.db_sel_default_cat()
    print(dan)
    print(callback.message.chat.id)
    cat = ""
    wr = ""
    for i in dan:
        cat = i[0]
        wr = i[1]
        await sqlite.db_insert_cat(cat,wr, callback.message.chat.id, i[2])

    await callback.message.edit_text(
        "Категории заполнены"
    )


# Переход категорий для редактирование
@form_router.callback_query(keyboards.keyboard.NumbersCallbackFactory.filter(F.action == "edit_cats"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.NumbersCallbackFactory
        ) -> None:
    name_cat = callback_data.name
    wiring = callback_data.val
    id_cat = callback_data.value
    wiring_e = ""
    wiring_n = ""
    if wiring == "decr":
        wiring_e = "(расход)"
        wiring_n = "приход"
    elif wiring == "incr":
        wiring_e = "(приход)"
        wiring_n = "расход"

    await callback.message.edit_text(
        f"Редактирование категории <b>{name_cat} {wiring_e}</b>",
         reply_markup=keyboards.keyboard.gen_edits_cat(id_cat,wiring_n,name_cat,wiring)
    )


# Запрос списка категорий для редактирование EXIT
@form_router.callback_query(keyboards.keyboard.NumbersCallbackFactory.filter(F.action == "edit_cats_ex"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
) -> None:

    await callback.message.edit_text("Список категорий\n---\n Выход")

# Удоляем категорию
@form_router.callback_query(keyboards.keyboard.NumbersCallbackFactory.filter(F.action == "edit_cats_del"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.NumbersCallbackFactory
) -> None:
    name_cat = callback_data.name
    id = callback_data.value
    await sqlite.db_del_edit_cat(id)
    await callback.message.edit_text(f'Категория "{name_cat}" \n---\n❌ Удалена')


# Назначаем категорию как топливную
@form_router.callback_query(keyboards.keyboard.NumbersCallbackFactory.filter(F.action == "edit_cats_fuel"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.NumbersCallbackFactory
) -> None:
    name_cat = callback_data.name
    id = callback_data.value
    await sqlite.db_upg_edit_cat(id,"yes","decr")
    await callback.message.edit_text(f"Назначили {name_cat} как топливную категорию  \n---\n")

# Изменяем в категории операцию
@form_router.callback_query(keyboards.keyboard.NumbersCallbackFactory.filter(F.action == "edit_cats_wiring"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.NumbersCallbackFactory
) -> None:

    name_cat = callback_data.name
    id = callback_data.value
    wiring = callback_data.val
    wiring_e = ""
    wiring_n = ""
    if wiring == "decr":
        wiring_e = "incr"
        wiring_n = "приходную"
    elif wiring == "incr":
        wiring_e = "decr"
        wiring_n = "расходную"

    await sqlite.db_upg_edit_cat_wr(id,wiring_e)
    await callback.message.edit_text(f'Назначили категорию: "{name_cat}" как {wiring_n}  \n---\n')


@form_router.message(Command("cattegor_incr"))
async def buy(message: types.Message):

    data = await sqlite.db_sel_inc()
    await message.answer(f"Список: {data}", reply_markup=keyboards.keyboard.gen_markup(data))


@form_router.callback_query(keyboards.keyboard.NumbersCallbackFactory.filter(F.action == "ins_cat"))
async def command_start(callback: types.CallbackQuery, state: FSMContext) -> None:
    await state.set_state(Form.cat_new)
    await callback.message.edit_text(
        "Введите название категории"
    )


@form_router.message(Form.cat_new)
async def process_name(message: Message, state: FSMContext) -> None:
    await state.update_data(cat_new=message.text)
    await state.update_data(id_us=message.from_user.id)

    await message.answer(
        f"Как будем учитывать категорию, {html.quote(message.text)} ?",
        reply_markup=get_and_cat()
    )
def get_and_cat():
    buttons = [
        [
            types.InlineKeyboardButton(text="Расход", callback_data="num_decr"),
            types.InlineKeyboardButton(text="Приход", callback_data="num_incr")
        ],

    ]
    keyboard = types.InlineKeyboardMarkup(inline_keyboard=buttons)
    return keyboard


@form_router.message(Command("flight"))
async def process_name(message: Message, state: FSMContext) -> None:
    dann = await sqlite.db_sel_flight("active",message.from_user.id)
    if not dann:
        dat = "🛣 Открыть"
        act = "fli_new"
        await message.answer(
            f"<b>Управление рейсами.</b> \nУ вас нет активных рейсов.",
                reply_markup=keyboards.keyboard.key_fl_config(dat,act)
                )

    else:
        dat = "🛣 Завершить"
        act = "fli_off"
        for i in dann:

            await message.answer(
                f"<b>Управление рейсами.</b> \nЗавершить рейс: <b>{i[1]} от {i[2]} </b>",
                reply_markup=keyboards.keyboard.key_fl_config_activ(dat,act, i[0])

            )
        print(dann)
        await state.update_data(user_id_fli=message.from_user.id)


@form_router.message(Command("list_cat"))
async def process_name(message: Message, state: FSMContext) -> None:
    dann = await sqlite.db_sel_flight("active",message.from_user.id)
    await message.answer(
        "List categor",
        reply_markup=keyboards.keyboard.gen_list_cat()
    )




############ Запрос списка рейсов
@form_router.callback_query(keyboards.keyboard.Flight_Callback.filter(F.action == "list_flights"))
async def callbacks_num_change_fab(callback: types.CallbackQuery) -> None:
    users = callback.message.chat.id
    data = await sqlite.db_sel_flight_name(users)

    await callback.message.edit_text(
        f"<b>Список последних 10 рейсов</b>",
        reply_markup=keyboards.keyboard.gen_list_flights(data)
    )


############ Запрос списка поступлений
@form_router.callback_query(keyboards.keyboard.Flight_Callback.filter(F.action == "list_receipt"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.Flight_Callback
) -> None:
    users = callback.message.chat.id
    id_fl = callback_data.id_fl
    data = await sqlite.db_sel_tr_pay(id_fl,"incr")

    await callback.message.edit_text(
        f"<b>Список поступлений в рейсе</b>",
        reply_markup=keyboards.keyboard.gen_list_pay(data)
    )

############ Запрос списка оплат
@form_router.callback_query(keyboards.keyboard.Flight_Callback.filter(F.action == "list_pay"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.Flight_Callback
) -> None:
    users = callback.message.chat.id
    id_fl = callback_data.id_fl

    data = await sqlite.db_sel_tr_pay(id_fl,"decr")

    await callback.message.edit_text(
        f"<b>Список оплат в рейсе</b>",
        reply_markup=keyboards.keyboard.gen_list_pay(data)
    )

############ Запрос списка заправок
@form_router.callback_query(keyboards.keyboard.Flight_Callback.filter(F.action == "list_fuels"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.Flight_Callback
        ) -> None:
    users = callback.message.chat.id
    id_fl = callback_data.id_fl
    data = await sqlite.db_sel_tr_fuels(id_fl)
    print(data)
    litr = 0
    if len(data) != 0:
        for i in data:
            litr = litr + i[2]

        await callback.message.edit_text(
            f"<b>⛽ Список заправок в рейсе</b>\nЗаправлено: <b>{litr}л.</b>",
            reply_markup=keyboards.keyboard.gen_list_fuels(data)
        )
    else:
        await callback.message.edit_text(f"<b>⛽ Учтённых заправок нет.</b>" )

############ В Рейс
@form_router.callback_query(keyboards.keyboard.Flight_Callback.filter(F.action == "fli_new"))
async def callbacks_num_change_fab(callback: types.CallbackQuery) -> None:
    data = await sqlite.db_sel_cars()
    await callback.message.edit_text(
        "Выбери транспорт",
        reply_markup=keyboards.keyboard.gen_select_cars(data)
        )



@form_router.callback_query(keyboards.keyboard.NumbersCallbackFactory.filter(F.action == "selcars"))
async def process_name_flight(callback: types.CallbackQuery,
                                  callback_data: keyboards.keyboard.NumbersCallbackFactory,
                                  state: FSMContext) -> None:
    users = callback.message.chat.id
    print(f"ID users: {users}")

    await state.update_data(carsM=callback_data.val)
    await state.update_data(cars=callback_data.value)
    data = await sqlite.db_sel_nam_point(users)
    await callback.message.edit_text(f"<b>Куда едим ?</b>\nВыбери из списка или введи новый пункт.", reply_markup=keyboards.keyboard.gen_select_point(data))
    await state.set_state(Flight.name)

############ Обработка Выбора Конечной точкт Рейса
@form_router.callback_query(keyboards.keyboard.Callback_Point.filter(F.action == "point"))
async def process_name_flight(callback: types.CallbackQuery,
                                  callback_data: keyboards.keyboard.Callback_Point,
                                  state: FSMContext) -> None:

    await state.update_data(name=callback_data.value)
    await callback.message.answer(
        text=f"Рейс : {callback_data.value} \nВведи начальный километраж.",
        reply_markup=ReplyKeyboardRemove(),
    )
    await state.set_state(Flight.mileage)


############ Обработка ввода Конечной точкт Рейса
@form_router.message(Flight.name)
async def process_name_flight(message: Message, state: FSMContext) -> None:
    await state.update_data(name=message.text)
    name = message.text
    await message.answer(
            text=f"Рейс : {name} \nВведи начальный километраж.",
        reply_markup=ReplyKeyboardRemove(),
    )
    await state.set_state(Flight.mileage)

############ Обработка ввода начального пробега
@form_router.message(Flight.mileage)
async def process_write_mileage(message: Message, state: FSMContext) -> None:
    await state.update_data(mileage=message.text)
    await message.answer(
        text=f"Введи остаток топлива.",
        reply_markup=ReplyKeyboardRemove(),
    )
    await state.set_state(Flight.fuel)

############ Обработка ввода начального остатка топлива
@form_router.message(Flight.fuel)
async def process_write_mileage(message: Message, state: FSMContext) -> None:
    await state.update_data(fuel=message.text)
    await state.update_data(id_us=message.from_user.id)
    data = await state.get_data()
    mileage = data["mileage"]
    name = data["name"]
    fuel_in = data["fuel"]
    cars = data["cars"]
    carsMarka = data["carsM"]
    await message.answer(
            text=f"Рейс : {name} от {transform_date(now1)}\nТранспорт: <b>{carsMarka}</b>\nНачальный километраж: <b>{mileage}</b>\n Остаток топлива: <b>{fuel_in}</b>\n Верно?",
        reply_markup=keyboards.keyboard.key_flight(),
    )


@form_router.callback_query(keyboards.keyboard.Flight_Callback.filter(F.action == "fli_okdist"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        state: FSMContext
) -> None:

    data = await state.get_data()
    mileage = data["mileage"]
    name = data["name"]
    fuel = data["fuel"]
    id_us = data["id_us"]
    cars = data["cars"]
    carsMarka = data["carsM"]
    status = "active"
    await callback.message.edit_text(
        f"записал:\n Рейс : {name} от {transform_date(now1)}\n Транспорт: <b>{carsMarka}</b>\nНачальный километраж: <b>{mileage}</b>\n Остаток топлива: {fuel}."
    )
    await sqlite.db_insert_flight(name,id_us,fuel,mileage,now,cars,status)
    await state.clear()



@form_router.callback_query(keyboards.keyboard.Flight_Callback.filter(F.action == "now_cars"))
async def process_name_flight(callback: types.CallbackQuery, state: FSMContext) -> None:

    await callback.message.edit_text("Отмена")
    await state.clear()


############/ В Рейс /##########################################################

############ Из Рейса /##########################################################

@form_router.callback_query(keyboards.keyboard.Flight_Callback.filter(F.action == "fli_off"))
async def process_name(callback: types.CallbackQuery, state: FSMContext) -> None:
    print(callback.message.from_user.id)


    dat = await state.get_data()
    user_id_fli = dat["user_id_fli"]

    dann = await sqlite.db_sel_flight("active", user_id_fli)

    for i in dann:

        await callback.message.edit_text(
            f"Завершить рейс: <b>{i[1]} от {i[2]} </b>",
            reply_markup=get_key_fl()
        )

        await state.update_data(name= i[1]+" от "+ i[2])
        await state.update_data(id = i[0])
        await state.update_data(date_in = i[2])
        await state.update_data(in_fuel=i[4])
        await state.update_data(in_odometr=i[5])
        await state.update_data(cars=i[3])
def get_key_fl():
    buttons = [
        [
            types.InlineKeyboardButton(text="Да", callback_data="exit_flight"),
            types.InlineKeyboardButton(text="Нет", callback_data="num_incr")
        ],
    ]
    keyboard = types.InlineKeyboardMarkup(inline_keyboard=buttons)
    return keyboard

@form_router.callback_query(F.data.startswith("exit_flight"))
async def callbacks_num(callback: types.CallbackQuery, state: FSMContext):
    await callback.message.edit_text("Введи конечный километраж.")
    await state.set_state(Flight_exit.finish_odometr)
    print("1111")


@form_router.message(Flight_exit.finish_odometr)
async def process_write_mileage(message: Message, state: FSMContext) -> None:
    print(message.text)
    await state.update_data(odometr=message.text)
    await message.answer(
        text=f"Введи остаток топлива.",
        reply_markup=ReplyKeyboardRemove(),
    )
    await state.set_state(Flight_exit.fuel_residues)

@form_router.message(Flight_exit.fuel_residues)
async def process_write_mileage(message: Message, state: FSMContext) -> None:
    ful = await state.update_data(fuel=message.text)
    dat = await state.get_data()
    id = dat["id"]
    carsMarka = dat["cars"]
    name = dat["name"]
    odometr = dat["odometr"]
    fuel = dat["fuel"]
    status = "fulfilled"
    #  await sqlite.db_obdate_flight(id,fuel,odometr,now,status)

    await message.answer(
        f"Закрываем рейс с данными?\n Рейс : {name} \n Транспорт: {carsMarka}\nКонечный километраж: {odometr}\n Остаток топлива: {fuel}.",
        reply_markup=keyboards.keyboard.key_flight_confirmation()
    )

@form_router.callback_query(keyboards.keyboard.Flight_Callback.filter(F.action == "fli_confirmation"))
async def process_like_write_bots(callback: types.CallbackQuery,
                                  callback_data: keyboards.keyboard.NumbersCallbackFactory,
                                  state: FSMContext) -> None:


    dat = await state.get_data()
    id = dat["id"]
    carsMarka = dat["cars"]
    name = dat["name"]

    in_odometr = dat["in_odometr"]
    odometr = dat["odometr"]
    Mileage_covered = int(odometr) - int(in_odometr)

    in_fuel = dat["in_fuel"]
    fuel = dat["fuel"]
    Fuel_used_up = int(in_fuel)-int(fuel)

    Fuel_consumption = round(Fuel_used_up / Mileage_covered * 100, 2)

    status = "fulfilled"
    await sqlite.db_obdate_flight(id,fuel,odometr,now,status)
    print(id)
    await exel_creat(id)

    await callback.message.edit_text(
        f"Записал. \nРейс : {name} Пройденный километраж: {Mileage_covered}\n Израсходовано топлива: {Fuel_used_up}\n Расход топлива: {Fuel_consumption}"
    )
    # await exel_creat(id)
    await state.clear()






async def update_num_text(message: types.Message, new_value: str):
    await message.edit_text(
        f"Уитывать категорию: {new_value} как",
        reply_markup=get_keyboard()
    )
def get_keyboard():
    buttons = [
        [
            types.InlineKeyboardButton(text="Расход", callback_data="num_decr"),
            types.InlineKeyboardButton(text="Приход", callback_data="num_incr")
        ],
        [types.InlineKeyboardButton(text="Подтвердить", callback_data="num_finish")]
    ]
    keyboard = types.InlineKeyboardMarkup(inline_keyboard=buttons)
    return keyboard



###################################################
@form_router.callback_query(F.data.startswith("num_"))
async def callbacks_num(callback: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    us_id = data["id_us"]
    cat_new = data["cat_new"]
    action = callback.data.split("_")[1]
    await state.clear()

    if action == "incr":
        await sqlite.db_insert_cat(cat_new, "incr", us_id," ")
        await callback.message.edit_text(f'Добавлена категория : <b>"{cat_new}"</b> как поступление.')

    elif action == "decr":
        await sqlite.db_insert_cat(cat_new, "decr", us_id,"")
        await callback.message.edit_text(f'Добавлена категория : <b>"{cat_new}"</b> как расход.')

    elif action == "finish":
        await sqlite.db_insert_cat(cat_new, "finish", us_id)
        await callback.message.edit_text(f"Итого: {us_id} finish {cat_new}")

    await callback.answer()


@form_router.message(Form.cat_new, F.text.casefold() == "No")
async def process_like_write_bots(message: Message, state: FSMContext) -> None:
    await state.clear()
    await message.reply(
        "Круто! Я тоже в Расход",
        reply_markup=ReplyKeyboardRemove(),
    )


# @form_router.message(Command("cancel"))
@form_router.message(F.text.casefold() == "cancel")
async def cancel_handler(message: Message, state: FSMContext) -> None:
    """
    Allow user to cancel any action
    """
    current_state = await state.get_state()
    if current_state is None:
        return

    logging.info("Cancelling state %r", current_state)
    await state.clear()
    await message.answer(
        "Cancelled.",
        reply_markup=ReplyKeyboardRemove(),
    )

###### ОТЧЁТ
# @form_router.message(Command("order"))
@form_router.callback_query(keyboards.keyboard.Flight_Callback.filter(F.action == "order_flights"))
async def cancel_handler(callback: types.CallbackQuery,
                         callback_data: keyboards.keyboard.Flight_Callback,
                         state: FSMContext) -> None:
    user_id = callback.message.chat.id
    flight_id = callback_data.id_fl
    status_fl = callback_data.st_fl

    fuel = 0
    fuels = 0
    # print(f"id {user_id} fl_id: {flight_id}")
    order_str = ""
    sum_incr = 0
    sum_decr = 0
    driver = ""
    order_str = order_str + "<b>Поступление:</b>\n"
   # Запросы
    incrdat = await sqlite.db_sel_tr_order(flight_id,"incr")
    decrdat = await sqlite.db_sel_tr_order(flight_id, "decr")
    odometr = await sqlite.db_order_info(flight_id)
    drivers = await sqlite.db_sel_driver(user_id)
    # print(f"drivers {drivers}")
    for i in drivers:
        driver = i[0]

    for i in incrdat:
        order_str = order_str + f"  {i[1]} Σ: {i[2]}р.\n"
        sum_incr = sum_incr + i[2]
    order_str = order_str + f"<b>Итого</b> получено: <b>{sum_incr}р.</b>\n"

    order_str = order_str + "\n<b>Расходы:</b>\n"
    for i in decrdat:
        order_str = order_str + f"  {i[1]} Σ: {i[2]}р.\n"
        sum_decr = sum_decr + i[2]
        if i[3] != 0.0:
            fuel = i[3]

    order_str = order_str + f"<b>Итого</b> расходов: <b>{sum_decr}р.</b>\n"
    order_str = order_str + f"<b>Остатк</b> ДС: <b>{sum_incr-sum_decr}р.</b>\n"
    order_str = order_str + f"<b>\nТопливо:</b>\n"
    # odom_in = 0
    # odom_ex = 0
    odom = 0
    auto =""
    for i in odometr:
        order_str = order_str + f"  Начальный остаток: {i[0]} л.\n  Заправлено: {fuel} л.\n"
        fuels = fuel + i[0]

        if i[1] != None:
             fuels = fuels - i[1]
             order_str = order_str + f"  Конечный остаток: {i[1]} л.\n<b>Израсходовано:</b> {fuels} л.\n\n<b>Одометр\n</b>"

        else:
            order_str = order_str + f"\n\n<b>Одометр\n</b>"

        order_str = order_str + f"  Начальный: {i[2]} km.\n"
        odom = i[2]

        if i[3] != None:
            # odom_ex = i[3]
            # odom = odom_ex - odom_in
            odom = abs(i[3] -odom)
            order_str = order_str + f"  Конечный: {i[3]} km.\n<b>Пройдено : {odom} km.</b>"
            order_str = order_str + f"\n \n<b>Средний расход:</b> {round(fuels / odom * 100,2)} л/100км"

        auto = i[6]
        order_str = order_str + f"\nАвто: {auto} {i[7]}"
        order_str = order_str + f"\nВодитель: {driver}"


    if status_fl =="fulfilled":
        await callback.message.edit_text(order_str,
            reply_markup=keyboards.keyboard.gen_key_exel(flight_id)
        )

    else:
        await callback.message.edit_text(order_str)







async def main():
    bot = Bot(token=TOKEN, parse_mode=ParseMode.HTML)

    dp = Dispatcher()
    dp.include_router(form_router)

    await dp.start_polling(bot,
                           on_startup=on_startup
                           )


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO,
                        stream=sys.stdout,
                        )
    asyncio.run(main())
