import sqlite3
from datetime import datetime, timedelta, date
import asyncio
import logging
import sys
import keyboards.keyboard
import sqlite
import emoji
import calendar
import openpyxl
###################PDF

from win32com import client
from openpyxl.worksheet import worksheet
from datetime import datetime
from aiogram.exceptions import TelegramBadRequest
from aiogram import types
from os import getenv
from re import Match
from magic_filter import RegexpMode
from aiogram import Bot, Dispatcher, F, Router, html
from aiogram.filters import Command, CommandStart
from aiogram.fsm.context import FSMContext
from contextlib import suppress
from aiogram.fsm.state import State, StatesGroup
from aiogram.enums import ParseMode, ChatAction

from aiogram.types import (
    KeyboardButton,
    Message,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
)

TOKEN = getenv("BOT_TOKEN")

form_router = Router()


class Form(StatesGroup):
    name = State()
    like_bots = State()
    language = State()
    cash = State()
    id_us = State()
    cat_new = State()
    mes_id = State()
    coments = State()
    fuel_vol = State()


class Dates(StatesGroup):
    day = State()
    month = State()
    yer = State()
    itog = State()


class Fuels(StatesGroup):
    fuel_vol = State()


class Flight(StatesGroup):
    name = State()
    status = State()
    cars = State()
    mileage = State()
    fuel = State()


class Flight_exit(StatesGroup):
    status = State()
    name = State()
    finish_odometr = State()
    fuel_residues = State()


class orders_gen(StatesGroup):
    namber = State()


class Delivery(StatesGroup):
    and_us = State()
    and_phone = State()
    sender = State()
    recipient = State()
    sender_city = State()
    recipient_city = State()
    photo = State()


now = datetime.now().strftime("%Y-%m-%d %H:%M")
now2 = datetime.now().strftime("%d.%m.%Y")
now1 = datetime.now().strftime("%Y-%m-%d %H:%M:%S")


# delivery - Доставки
# categories - Управление категориями
# flight - Управление Рейсами
# cancel - Выход


async def on_startup(_):
    await sqlite.db_connect()
    print("Подключение БД ОК")


def transform_date(date):
    months = ['января', 'февраля', 'марта', 'апреля', 'мая', 'июня',
              'июля', 'августа', 'сентября', 'октября', 'ноября', 'декабря']
    day, month, year = now2.split('.')
    return f'{day} {months[int(month) - 1]} {year} года'


def transform_date_new(date):
    months = ['января', 'февраля', 'марта', 'апреля', 'мая', 'июня',
              'июля', 'августа', 'сентября', 'октября', 'ноября', 'декабря']
    year, month, day = date.split('-')
    return f'{day} {months[int(month) - 1]} {year} года'


def is_number(string):
    try:
        float(string)
        return True
    except ValueError:
        return False


@form_router.callback_query(keyboards.keyboard.Flight_Callback.filter(F.action == "order_flights_pdf"))
@form_router.message(Command("pdf"))
async def handle_command_pic(callback: types.CallbackQuery):
    # await message.bot.send_chat_action(
    #     chat_id=message.chat.id,
    #     action=ChatAction.UPLOAD_PHOTO,
    # )
    # Create a Workbook object
    await callback.message.answer("Создаю PDF")

    #
    # # Opening Microsoft Excel
    WB_PATH = r'test.xlsx'
    WB_PATH2 = "C:\\Users\\Сергей\\PycharmProjects\\pythonProject2\\test.xlsx"
    WB_PATH3 = "C:\\Users\\Сергей\\PycharmProjects\\pythonProject2\\test.pdf"
    file_dsc = "test.pdf"
    excel = client.Dispatch("Excel.Application")
    # Read Excel File
    sheets = excel.Workbooks.Open(WB_PATH2)
    work_sheets = sheets.Worksheets[0]
    # # Converting into PDF File
    work_sheets.ExportAsFixedFormat(0, WB_PATH3)
    sheets.Close(False)
    excel.Quit()

    await callback.message.reply_document(
        document=types.FSInputFile(
            path=file_dsc,
        ),
    )


@form_router.message(Command("pic"))
async def handle_command_pic(message: types.Message):
    await message.bot.send_chat_action(
        chat_id=message.chat.id,
        action=ChatAction.UPLOAD_PHOTO,
    )
    # url = "https://t4.ftcdn.net/jpg/00/97/58/97/360_F_97589769_t45CqXyzjz0KXwoBZT9PRaWGHRk5hQqQ.jpg"
    # url = "https://images.unsplash.com/photo-1608848461950-0fe51dfc41cb"
    file_path = "cat.jpg"
    file_dsc = "test.pdf"
    # await message.bot.send_photo()
    # await message.reply_photo(
    #     # photo=url,
    #     photo=types.FSInputFile(
    #         path=file_path,
    #         # filename=
    #     ),
    #     caption="cat small pic",
    # )
    await message.reply_document(
        document=types.FSInputFile(
            path=file_dsc,
        ),
    )


###########################
@form_router.callback_query(keyboards.keyboard.Div_Call.filter(F.action == "dev_add"))
async def callbacks_num(callback: types.CallbackQuery, state: FSMContext):
    await callback.message.edit_text("Введи номер или Фамилию отправителя")
    await state.set_state(Delivery.sender)


@form_router.message(Delivery.sender)
async def process_write_mileage(message: Message, state: FSMContext) -> None:
    print(message.text)
    text = message.text
    if is_number(text):

        data = await sqlite.db_dev_user_ph(text)
        if len(data) == 0:
            await message.reply(
                f"Контакт по номеру <b>{text}</b> не найден, Рекомендуется проверить данные если нужно исправить их или создать запись по кнопке.",
                reply_markup=keyboards.keyboard.get_dev_us_and("nul", text, "sender")
            )
            # await state.clear()
            await state.update_data(phone=text)
            # await state.set_state(Delivery.sender)
        else:
            for i in data:
                print(i[1])
            await message.reply(
                f"Найдены :",
                reply_markup=keyboards.keyboard.get_dev_us_sel(data, "senders")
            )

    else:

        data = await sqlite.db_dev_user_name(f"%{text.capitalize()}%")
        print(data)
        if len(data) == 0:
            await message.reply(
                f"Контакт по ФИО <b>{text}</b> не найден, Рекомендуется проверить данные если нужно исправить их или создать запись по кнопке.",
                reply_markup=keyboards.keyboard.get_dev_us_and(text, "nul", "sender")
            )
            await state.update_data(name=text)
            await state.set_state(Delivery.sender)
        else:
            for i in data:
                print(i[1])
            await message.reply(
                f"Найдены:",
                reply_markup=keyboards.keyboard.get_dev_us_sel(data, "senders")
            )


@form_router.callback_query(keyboards.keyboard.NumbersCallbackFactory.filter(F.action == "us_dev_and_phone"))
async def process_name_flight(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.NumbersCallbackFactory,
        state: FSMContext) -> None:
    await state.update_data(sen_rec=callback_data.val)

    data = await state.get_data()
    name = data["name"]
    await callback.message.edit_text(f'Наименование {name}.\nВведите номер в формате "+71231234567"')
    await state.set_state(Delivery.and_phone)


@form_router.callback_query(keyboards.keyboard.NumbersCallbackFactory.filter(F.action == "us_dev_and_name"))
async def process_name_flight(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.NumbersCallbackFactory,
        state: FSMContext) -> None:
    await state.update_data(sen_rec=callback_data.val)
    sen_rec = callback_data.val
    data = await state.get_data()
    phone = data["phone"]

    if sen_rec == "sender":

        await callback.message.edit_text(f'Номер тел. {phone}.\nВведите ФИО Отпровителя')
        await state.set_state(Delivery.and_us)

    elif sen_rec == "recipient":

        await callback.message.edit_text(f'Номер тел. {phone}.\nВведите ФИО Получателя')
        await state.set_state(Delivery.and_us)


@form_router.message(Delivery.and_us)
async def process_write_mileage(message: Message, state: FSMContext) -> None:
    name = message.text
    print(f"name {name}")
    data = await state.get_data()
    phone = data["phone"]
    sen_rec = data["sen_rec"]

    await sqlite.db_dev_user_insert_us(phone, name)

    if sen_rec == "sender":
        await message.answer(f'Добавлен отпровитель с данными:\nНомер тел.: {phone}.\nФИО: {name}')
        await state.set_state(Delivery.sender_city)
        await message.answer(f'Введите город отправления или выбрать')

        dat = await sqlite.db_dev_sel_id_user(name, phone)
        for i in dat:
            id = i[0]
            name_s = i[1]

            await state.update_data(senders_id=id)
            await state.update_data(senders_name=name_s)

    elif sen_rec == "recipient":
        await message.answer(f'Добавлен получатель с данными:\nНомер тел.: {phone}.\nФИО: {name}')
        await message.answer(f'Ввести город получения или выбрать')

        await state.set_state(Delivery.recipient_city)

        dat = await sqlite.db_dev_sel_id_user(name, phone)
        for i in dat:
            id = i[0]
            name_s = i[1]

            await state.update_data(recipient_id=id)
            await state.update_data(recipient_name=name_s)


@form_router.message(Delivery.and_phone)
async def process_write_mileage(message: Message, state: FSMContext) -> None:
    phone = message.text
    data = await state.get_data()
    name = data["name"]
    sen_rec = data["sen_rec"]

    await sqlite.db_dev_user_insert_us(phone, name)
    dat = await sqlite.db_dev_sel_id_user(name, phone)
    id = 0
    name_s = ""
    for i in dat:
        id = i[0]
        name_s = i[1]

    if sen_rec == "sender":
        await message.answer(f'Добавлен отпровитель с данными:\nНомер тел.: {phone}.\nФИО: {name}')
        await state.set_state(Delivery.sender_city)
        await message.answer(f'Введите город отправления или выбрать')
        await state.update_data(senders_id=id)
        await state.update_data(senders_name=name_s)


    elif sen_rec == "recipient":
        await message.answer(f'Добавлен получатель с данными:\nНомер тел.: {phone}.\nФИО: {name}')
        await state.set_state(Delivery.recipient_city)
        await message.answer(f'Ввести город получения или выбрать')
        await state.update_data(recipient_id=id)
        await state.update_data(recipient_name=name_s)


@form_router.message(Delivery.sender_city)
async def process_write_mileage(message: Message, state: FSMContext) -> None:
    await state.update_data(sender_city=message.text)
    await message.answer(f'Введи номер или ФИО получателя')
    await state.set_state(Delivery.recipient)


@form_router.message(Delivery.recipient)
async def process_write_mileage(message: Message, state: FSMContext) -> None:
    recipient = message.text
    print(message.text)
    if is_number(recipient):
        data = await sqlite.db_dev_user_ph(recipient)
        if len(data) == 0:
            await message.reply(
                f"Контакт по номеру <b>{recipient}</b> не найден, Рекомендуется проверить данные если нужно исправить их или создать запись по кнопке.",
                reply_markup=keyboards.keyboard.get_dev_us_and("nul", recipient, "recipient")
            )
            await state.update_data(phone=recipient)
        else:
            for i in data:
                print(i[1])
            await message.reply(
                f"Найдены:",
                reply_markup=keyboards.keyboard.get_dev_us_sel2(data, "recipient")
            )
    else:

        data = await sqlite.db_dev_user_name(f"%{recipient.capitalize()}%")
        print(data)
        if len(data) == 0:
            await message.reply(
                f"Контакт по ФИО <b>{recipient}</b> не найден, Рекомендуется проверить данные если нужно исправить их или создать запись по кнопке.",
                reply_markup=keyboards.keyboard.get_dev_us_and(recipient, "nul", "recipient")
            )
            await state.update_data(name=recipient)

        elif len(data) != 0:
            # for i in data:
            #     print(i[1])
            await message.reply(
                f"Найдены:",
                reply_markup=keyboards.keyboard.get_dev_us_sel2(data, "recipient")
            )
            # await state.set_state(Delivery.sender_city)


@form_router.callback_query(keyboards.keyboard.Div_Call.filter(F.action == "recipient"))
async def process_name_flight(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.Div_Call,
        state: FSMContext) -> None:
    recipient_id = callback_data.div_id
    recipient_name = callback_data.div_str
    await state.update_data(recipient_id=recipient_id)
    await state.update_data(recipient_name=recipient_name)
    await callback.message.edit_text(f'Введите город получения или выбрать')
    await state.set_state(Delivery.recipient_city)


@form_router.message(Delivery.recipient_city)
async def process_write_mileage(message: Message, state: FSMContext) -> None:
    recipient_city = message.text
    await state.update_data(recipient_city=recipient_city)
    data = await state.get_data()
    senders_name = data["senders_name"]
    sender_city = data["sender_city"]
    recipient_name = data["recipient_name"]
    recipient_city = data["recipient_city"]
    print(data)
    await message.answer(
        f'Отпровитель: {senders_name} из {sender_city} \nПолучатель: {recipient_name} в {recipient_city}\n Офрмить ?',
        reply_markup=keyboards.keyboard.get_dev_fin()
    )


@form_router.callback_query(keyboards.keyboard.NumbersCallbackFactory.filter(F.action == "devi_insert"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        state: FSMContext) -> None:
    data = await state.get_data()
    senders_name = data["senders_name"]
    sender_city = data["sender_city"]
    recipient_name = data["recipient_name"]
    recipient_city = data["recipient_city"]
    senders_id = data["senders_id"]
    recipient_id = data["recipient_id"]
    id = await sqlite.db_dev_insert(senders_id, sender_city, now1, recipient_id, recipient_city, "Принят")
    #
    # text_ms = f"Получатель {recipient_name} в {recipient_city}"

    await callback.message.edit_text(
        f"Добавлено №: {id} \n "
        f"Отпровитель: {senders_name} из {sender_city} \n"
        f"Получатель: {recipient_name} в {recipient_city}\n",
        reply_markup=keyboards.keyboard.get_keyboard_dev(id, callback.message.message_id)
    )
    await state.clear()


@form_router.callback_query(keyboards.keyboard.NumbersCallbackFactory.filter(F.action == "senders"))
async def process_name_flight(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.NumbersCallbackFactory,
        state: FSMContext) -> None:
    # await sqlite
    senders_id = callback_data.value
    senders_name = callback_data.name
    await state.update_data(senders_name=senders_name)
    await state.update_data(senders_id=senders_id)
    await callback.message.edit_text(f'Введите город отправления или выбрать')
    await state.set_state(Delivery.sender_city)


@form_router.callback_query(keyboards.keyboard.Div_Call.filter(F.action == "dev_end"))
async def process_name_flight(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.Div_Call,
        state: FSMContext) -> None:
    data = await sqlite.db_dev_for_issuance()
    await callback.message.edit_text(
        "Выбери отправления:",
        reply_markup=keyboards.keyboard.get_dev_delivir_sel(data)
    )
    print(f"data {data}")


## Перейти в посылку
@form_router.callback_query(keyboards.keyboard.Div_Call.filter(F.action == "open_pack"))
async def process_name_flight(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.Div_Call,
        state: FSMContext) -> None:
    # await handle_command_pic(callback)

    id_pac = callback_data.div_id
    photo = 0
    data = await sqlite.db_sel_pack(id_pac)
    data_foto = await sqlite.db_sel_pack_foto(id_pac)
    if len(data_foto) != 0:
        photo = 1
    else:
        photo = 0

    text_ms = ""
    id_photo = ""

    for i in data:
        print(i[0])
        # text_ms =f" из {i[0]} от \n{i[1]} \nв {i[3]} для <b>{i[2]}</b>\nДата отправления: {i[5]}\n📱  Телефон: \nК оплатте: <b>{i[4]}</b>"
        text_ms = f"\n🏠 Отпровитель: {i[0]} от \n{i[1]} \n📱  Телефон: {i[7]}\n\nПолучатель: <b>{i[2]}</b>\n📱  Телефон: {i[8]}\n🏠 Адрес: <b>{i[3]}</b>\n\nДата отправления: {i[5]}\n\nК оплатте: <b>{i[4]}</b>"

    await callback.message.edit_text(
        f"Отправление №{id_pac}\n {text_ms}",
        reply_markup=keyboards.keyboard.get_pac_dev(id_pac, photo)
    )


@form_router.callback_query(keyboards.keyboard.Div_Call.filter(F.action == "open_pack_foto"))
async def process_name_flight(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.Div_Call,
        state: FSMContext) -> None:
    # await handle_command_pic(callback)

    id_pac = callback_data.div_id

    data = await sqlite.db_sel_pack(id_pac)
    data_foto = await sqlite.db_sel_pack_foto(id_pac)

    ph1 = ()
    id_photo = ""
    text_ms = ""
    photos = []
    for ii in data_foto:
        photos.append(types.InputMediaPhoto(type='photo', media=ii[0], caption=ii[1]))

    await callback.message.reply_media_group(
        media=photos
    )  # Отправка фото

    # f"foto/{id}.jpg"
    #
    # for i in data:
    #     # print(i[0])
    #     id_photo = f"{i[6]}"
    #     print(i[6])
    #     text_ms = f" из {i[0]} от \n{i[1]} \nв {i[3]} для <b>{i[2]}</b>\nДата отправления: {i[5]}\nК оплатте: <b>{i[4]}</b>"

    # if id_photo != None:
    #     # await callback.message.reply_photo(
    #     #     photo=id_photo,
    #     #     caption=f"{text_ms}",
    #     #     reply_markup=keyboards.keyboard.get_pac_dev(id_pac,1)
    #     # )
    #     # await callback.message.reply_media_group(media=data_foto)
    #     await callback.message.edit_text(
    #         f"Фото отсутствует.\nОтправление №{id_pac} {text_ms}",
    #         reply_markup=keyboards.keyboard.get_pac_dev(id_pac)
    #     )
    #
    # else:
    #     await callback.message.edit_text(
    #         f"Фото отсутствует.\nОтправление №{id_pac} {text_ms}",
    #         reply_markup=keyboards.keyboard.get_pac_dev(id_pac)
    #     )


@form_router.callback_query(keyboards.keyboard.Div_Call.filter(F.action == "pack_issue"))
async def process_name_flight(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.Div_Call,
        state: FSMContext) -> None:
    id_pac = callback_data.div_id

    await callback.message.edit_text(
        "Вы действительно хотите выдать посылку ?",
        reply_markup=keyboards.keyboard.get_pac_dev_qu(id_pac)
    )

    # await state.set_state(Delivery.photo)


@form_router.callback_query(keyboards.keyboard.Div_Call.filter(F.action == "pack_issue1"))
async def process_name_flight(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.Div_Call,
        state: FSMContext) -> None:
    id_pac = callback_data.div_id

    await sqlite.db_issue_pack(id_pac, "доставлен")

    await callback.message.edit_text(
        f'Статус отправления №{id_pac} изменен на "Выдано"'
    )


@form_router.callback_query(keyboards.keyboard.Div_Call.filter(F.action == "dev_photo"))
async def process_name_flight(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.Div_Call,
        state: FSMContext) -> None:
    id_pac = callback_data.div_id
    await state.update_data(id_pac=id_pac)
    await callback.message.edit_text("Отправь фото")
    await state.set_state(Delivery.photo)


# # Обработка доставок
@form_router.message(Command("ph"))
async def buy(message: types.Message, state: FSMContext):
    await message.answer(
        f"Пришли мне фото",
        reply_markup=keyboards.keyboard.get_dev_pan()
    )

    await state.set_state(Delivery.photo)
    # await  keyboards.keyboard.Div_Call.action = "sdsds"


@form_router.message(Delivery.photo)
async def process_write_mileage(message: Message, state: FSMContext) -> None:
    us_id = "555115"
    id = message.photo[0].file_id
    # print(message.photo[0].file_id)

    data = await state.get_data()
    id_pac = data["id_pac"]
    user_id = message.from_user.id
    await sqlite.db_isrt_foto(id, id_pac, user_id, now1)

    file_name = f"foto/{id}.jpg"
    await message.bot.download(file=message.photo[-1].file_id, destination=file_name)
    await message.answer(
        "Фото загружено,\n Выбери дальнейшее действие.",
        reply_markup=keyboards.keyboard.get_pac_dev_qu_foto(id_pac)
    )


@form_router.callback_query(keyboards.keyboard.Div_Call.filter(F.action == "del_tr_deliv"))
async def process_name_flight(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.Div_Call,
        state: FSMContext) -> None:
    id = callback_data.div_id
    mes_txt = f"Отправление №:{id} из: "
    mes = await sqlite.db_dev_delliv(id)
    for i in mes:
        mes_txt = mes_txt + i[0]
        mes_txt = mes_txt + " от " + i[1]
        mes_txt = mes_txt + " в " + i[2]
        mes_txt = mes_txt + " для " + i[3]

    print(f"mes: {mes} mes_txt: {mes_txt}")
    await callback.message.edit_text(f"🚫 Отменено\n---\n{mes_txt} \n---\n")


###########################


############ Нажатие кнопки Создать Эксель
@form_router.callback_query(keyboards.keyboard.Flight_Callback.filter(F.action == "order_flights_exel"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.Flight_Callback
) -> None:
    flight_id = callback_data.id_fl

    await exel_creat(flight_id)
    await callback.message.edit_text("Создан отчёт")
    await callback.answer()


# # Обработка доставок
@form_router.message(Command("delivery"))
async def buy(message: types.Message, state: FSMContext):
    await message.answer(
        f"Управление доставками",
        reply_markup=keyboards.keyboard.get_dev_pan()
    )


# # Обработка Эксель
@form_router.message(Command("exel"))
async def buy(message: types.Message, state: FSMContext):
    await state.set_state(orders_gen.namber)
    await message.answer("Введи номер рейса")


@form_router.message(orders_gen.namber)
async def get_info_m(message: types.Message, state: FSMContext) -> None:
    id = message.text
    ppyt = "C:\\Users\\Сергей\\PycharmProjects\\pythonProject2\\Отчеты\\"

    data = await exel_creat(id)
    ppyt = ppyt + data[7:]
    await state.clear()

    await message.bot.send_chat_action(
        chat_id=message.chat.id,
        action=ChatAction.UPLOAD_DOCUMENT,
    )
    # await message.answer(f"{data}")

    WB_PATH3 = "C:\\Users\\Сергей\\PycharmProjects\\pythonProject2\\test.pdf"

    file_dsc = "test.pdf"
    excel = client.Dispatch("Excel.Application")
    # # Read Excel File
    sheets = excel.Workbooks.Open(ppyt)
    work_sheets = sheets.Worksheets[0]
    # # # Converting into PDF File
    work_sheets.ExportAsFixedFormat(0, WB_PATH3)
    sheets.Close(False)
    excel.Quit()

    await message.reply_document(
        document=types.FSInputFile(
            path=file_dsc,
            filename="Отчёт.pdf"
        ),
    )


# Обработка Эксель
async def exel_creat(id_fl: int):
    filename = "templete.xlsx"
    book = openpyxl.load_workbook(filename=filename)
    sheet: worksheet = book.worksheets[0]

    incr_db = await sqlite.db_sel_other_shapka(id_fl)
    auto = ""
    name = ""
    for i in incr_db:
        sheet["B1"].value = i[0]
        name = i[1]
        auto = i[1]

        sheet["B3"].value = auto + " " + i[2]
        sheet["B4"].value = i[3]

        sheet["J2"].value = i[4]
        date_in = datetime.strptime((i[4]), '%Y-%m-%d %H:%M')
        date_ex = datetime.strptime((i[5]), '%Y-%m-%d %H:%M')
        sheet["L2"].value = i[5]
        s = i[5]

        # print(s[:-6])
        name = name + "От" + s[:-6]
        sheet["J3"].value = i[6]
        sheet["L3"].value = i[7]
        sheet["J4"].value = i[8]
        sheet["L4"].value = i[9]
        sheet["E4"].value = i[10]
        delta_days = (date_ex - date_in).days
        sheet["N2"].value = delta_days

    incr_db = await sqlite.db_sel_tr_pay(id_fl, "incr")
    fuel_db = await sqlite.db_sel_tr_fuels(id_fl)
    other_db = await sqlite.db_sel_tr_other(id_fl)

    # # Заполняем поступление
    #     print(incr_db)
    nomer = 7
    for i in incr_db:
        pay = i[1]

        if 'карт' in pay:
            sheet.cell(row=nomer, column=2).value = i[2]
        else:
            sheet.cell(row=nomer, column=3).value = i[2]

        date = datetime.strptime((i[3]), '%Y-%m-%d %H:%M:%S')
        sheet.cell(row=nomer, column=1).value = date.strftime('%d-%m-%y %H:%M')
        nomer = nomer + 1

    # Заполняем топливо
    nomer = 7
    for i in fuel_db:
        date = datetime.strptime((i[1]), '%Y-%m-%d %H:%M:%S')
        sheet.cell(row=nomer, column=6).value = date.strftime('%d-%m-%y %H:%M')
        sheet.cell(row=nomer, column=7).value = i[2]
        sheet.cell(row=nomer, column=8).value = i[3]
        nomer = nomer + 1

    # Заполняем прочие расходы
    nomer = 7
    coll = 0
    categ = ""

    for i in other_db:

        if nomer > 24:
            nomer = 7
            coll = 3

        date = datetime.strptime((i[1]), '%Y-%m-%d %H:%M:%S')
        sheet.cell(row=nomer, column=9 + coll).value = value = date.strftime('%d-%m-%y %H:%M')
        categ = emoji.replace_emoji(i[2], replace='')
        sheet.cell(row=nomer, column=11 + coll).value = i[3]

        if i[4] != None:
            categ = categ + " " + f'"{emoji.replace_emoji(i[4], replace="")}"'
            sheet.cell(row=nomer, column=10 + coll).value = categ
        else:
            sheet.cell(row=nomer, column=10 + coll).value = categ
        nomer = nomer + 1

    filepath = f'Отчеты/{name}.xlsx'

    book.save(filepath)
    book.close()

    # excel_app = client.DispatchEx("Excel.Application")
    #
    # workbook = excel_app.Workbooks.Open(filepath)
    # worksheet = workbook.Worksheets(1)
    #
    # pdf_path = "output_file.pdf"
    #
    # worksheet.ExportAsFixedFormat(0, pdf_path)
    # workbook.Close(False)
    # excel_app.Quit()
    #
    return filepath
    # # file_dsc = pdf_path
    #


def date_diff(date1, date2):
    d1 = date.strptime(date1, '%Y-%m-%d')
    t1 = datetime.strptime(date1 + ' ' + date1.split()[-1], '%Y-%m-%d %H:%M').time()
    d2 = date.strptime(date2, '%Y-%m-%d')
    t2 = datetime.strptime(date2 + ' ' + date2.split()[-1], '%Y-%m-%d %H:%M').time()

    delta = d2 - d1 - timedelta(hours=t2.hour - t1.hour, minutes=t2.minute - t1.minute)
    return delta.days


# Обработка выбора меню категории
@form_router.callback_query(keyboards.keyboard.NumbersCallbackFactory.filter(F.action == "select"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.NumbersCallbackFactory,
        state: FSMContext
) -> None:
    await state.update_data(name=callback_data.val)
    data = await state.update_data(language=callback_data.value)
    await state.update_data(mes_id=callback.message.message_id)

    cash = data["cash"]
    sell = data["language"]
    cat = data["name"]
    us_id = data["id_us"]
    flight = data["id_flight"]

    if sell == 123:

        await callback.message.edit_text(
            "💬 Введите, количество литров."
        )
        await state.set_state(Form.fuel_vol)
        await state.update_data(mes_text=callback.message)


    else:
        await select_cat(callback.message, cat, sell, cash, us_id, 0, flight, "")
        await state.clear()
        print("mes_ID")
        print(callback.message.message_id)

    await callback.answer()


@form_router.message(Form.fuel_vol)
async def get_info_m(message: types.Message, state: FSMContext) -> None:
    litr = message.text
    litr = litr.replace(",", ".")
    try:
        float(litr)
        data = await state.get_data()
        cash = data["cash"]
        sell = data["language"]
        cat = data["name"]
        us_id = data["id_us"]
        flight = data["id_flight"]
        mes_text = data["mes_text"]

        await select_cat(mes_text, cat, sell, cash, us_id, 0, flight, litr)
        await state.clear()

    except ValueError:
        await state.set_state(Form.fuel_vol)
        await message.reply(
            f'Введённое значение не является числом!\n💬 Введите, количество литров. в формате "112.5"'
        )


# Обработка del trancaction
@form_router.callback_query(keyboards.keyboard.Itog_Callback.filter(F.action == "tr_cancel"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.Itog_Callback
) -> None:
    id_tranz = callback_data.id_tranz
    user = callback_data.val
    data = await sqlite.sel_tranc(id_tranz, user)

    cash = ""
    cat = ""
    date = ""
    for i in data:
        cash = i[0]
        cat = i[1]
        date = i[2]

    await sqlite.db_del_transactions(id_tranz, user)
    await callback.message.edit_text(
        f"Добавлено <b>{cat}</b> ₽ в категорию <b>{cash}</b> операция расход \n {date} \n---\n🚫 Отменено"
    )


# Обработка редактирование даты
@form_router.callback_query(keyboards.keyboard.Itog_Callback.filter(F.action == "tr_date"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.Itog_Callback,
        state: FSMContext
) -> None:
    # mes_id = mes_id,
    id_tranz = callback_data.id_tranz
    us_id = callback_data.val
    # mes_id = callback.data.mes_id

    data = await sqlite.sel_tranc(id_tranz, us_id)

    cash = ""
    cat = ""
    date = ""
    for i in data:
        cash = i[0]
        cat = i[1]
        date = i[2]

    dates = datetime.strptime(date, '%Y-%m-%d %H:%M:%S')
    ss = dates.strftime('%d-%m-%y %H:%M')
    month = dates.month
    year = dates.year
    days_in_month = calendar.monthrange(dates.year, dates.month)[1]  # 28
    text_mes = f"Добавлено <b>{cat}</b> ₽ в категорию <b>{cash}</b> операция расход \n {date} "
    await state.update_data(text_mes=text_mes)
    await state.update_data(dates=date)
    await state.update_data(dates_in=date)
    await state.update_data(id_tranz=id_tranz)
    await state.update_data(us_id=us_id)

    await callback.message.edit_text(
        text_mes + "\n<b>Изменить дату на:</b>",
        reply_markup=keyboards.keyboard.get_key_date(id_tranz, dates.day, month, year)
    )


# Обработка нопки назад к дате
@form_router.callback_query(keyboards.keyboard.Itog_Callback.filter(F.action == "pr_date"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.Itog_Callback,
        state: FSMContext
) -> None:
    data = await state.get_data()
    date = data["dates"]
    id_tranz = data["id_tranz"]
    text_mes = data["text_mes"]
    dates = datetime.strptime(date, '%Y-%m-%d %H:%M:%S')

    await callback.message.edit_text(
        text_mes + "\n<b>Изменить дату на:</b>",
        reply_markup=keyboards.keyboard.get_key_date(id_tranz, dates.day, dates.month, dates.year)
    )


###################################################
# Обработка кнопки готово дата
@form_router.callback_query(keyboards.keyboard.Itog_Callback.filter(F.action == "dates_finish"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        state: FSMContext
) -> None:
    data = await state.get_data()
    text_mes = data["text_mes"]
    id_tranz = data["id_tranz"]
    date = data["dates"]
    date_in = data["dates_in"]
    us_id = data["us_id"]

    dates = datetime.strptime(date, '%Y-%m-%d %H:%M:%S')
    dates_in = datetime.strptime(date_in, '%Y-%m-%d %H:%M:%S')

    if dates == dates_in:

        text_mes = text_mes[:-20] + f" {transform_date_new(date[:-9])} {date[11:]}"
    else:
        await sqlite.update_trans_date(date, id_tranz)
        text_mes = text_mes[:-20] + f"Дата изменена на {transform_date_new(date[:-9])} {date_in[11:]}"

    await callback.message.edit_text(
        text_mes,
        reply_markup=keyboards.keyboard.get_keyboard(id_tranz, callback.message.message_id, us_id)
    )
    await state.clear()


###################################################
# Обработка редактирование года
@form_router.callback_query(keyboards.keyboard.Itog_Callback.filter(F.action == "dates_year"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        state: FSMContext
) -> None:
    data = await state.get_data()
    text_mes = data["text_mes"]
    id_tranz = data["id_tranz"]
    date = data["dates"]
    us_id = data["us_id"]

    dates = datetime.strptime(date, '%Y-%m-%d %H:%M:%S')

    await callback.message.edit_text(
        text_mes + "\n<b>Выбери год:</b>",
        reply_markup=keyboards.keyboard.get_key_year(dates.year, id_tranz, us_id)
    )


# Установка выбраного года
@form_router.callback_query(keyboards.keyboard.Itog_Callback.filter(F.action == "year_set"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.Itog_Callback,
        state: FSMContext
) -> None:
    data = await state.get_data()
    date = data["dates"]
    id_tranz = data["id_tranz"]
    text_mes = data["text_mes"]
    dates = datetime.strptime(date, '%Y-%m-%d %H:%M:%S')
    select_year = callback_data.val
    d_new = dates.replace(year=select_year)
    await state.update_data(dates=d_new.strftime('%Y-%m-%d %H:%M:%S'))

    await callback.message.edit_text(
        text_mes + "\n<b>Изменить дату на:</b>",
        reply_markup=keyboards.keyboard.get_key_date(id_tranz, d_new.day, d_new.month, d_new.year)
    )


###################################################
# Обработка редактирование месяца
@form_router.callback_query(keyboards.keyboard.Itog_Callback.filter(F.action == "dates_month"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        state: FSMContext
) -> None:
    data = await state.get_data()
    text_mes = data["text_mes"]
    id_tranz = data["id_tranz"]
    us_id = data["us_id"]

    await callback.message.edit_text(
        text_mes + "\n<b>Выбери месяц:</b>",
        reply_markup=keyboards.keyboard.get_key_month(id_tranz, us_id)
    )


# Установка выбраного месяца
@form_router.callback_query(keyboards.keyboard.Itog_Callback.filter(F.action == "month_set"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.Itog_Callback,
        state: FSMContext
) -> None:
    data = await state.get_data()
    date = data["dates"]
    id_tranz = data["id_tranz"]
    text_mes = data["text_mes"]
    dates = datetime.strptime(date, '%Y-%m-%d %H:%M:%S')
    select_month = callback_data.val
    d_new = dates.replace(month=select_month)
    await state.update_data(dates=d_new.strftime('%Y-%m-%d %H:%M:%S'))

    await callback.message.edit_text(
        text_mes + "\n<b>Изменить дату на:</b>",
        reply_markup=keyboards.keyboard.get_key_date(id_tranz, d_new.day, d_new.month, d_new.year)
    )


# Обработка редактирование даты
@form_router.callback_query(keyboards.keyboard.Itog_Callback.filter(F.action == "dates_days"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        state: FSMContext
) -> None:
    data = await state.get_data()
    text_mes = data["text_mes"]
    date = data["dates"]
    id_tranz = data["id_tranz"]
    us_id = data["us_id"]
    dates = datetime.strptime(date, '%Y-%m-%d %H:%M:%S')
    days_in_month = calendar.monthrange(dates.year, dates.month)[1]  # 28
    await callback.message.edit_text(
        text_mes + "\n<b>Выбери день:</b>",
        reply_markup=keyboards.keyboard.get_key_days(days_in_month, id_tranz, us_id)
    )


# Установка выбраного дня
@form_router.callback_query(keyboards.keyboard.Itog_Callback.filter(F.action == "days_set"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.Itog_Callback,
        state: FSMContext
) -> None:
    data = await state.get_data()
    date = data["dates"]
    id_tranz = data["id_tranz"]
    text_mes = data["text_mes"]
    dates = datetime.strptime(date, '%Y-%m-%d %H:%M:%S')
    select_day = callback_data.val
    d_new = dates.replace(day=select_day)
    await state.update_data(dates=d_new.strftime('%Y-%m-%d %H:%M:%S'))

    await callback.message.edit_text(
        text_mes + "\n<b>Изменить дату на:</b>",
        reply_markup=keyboards.keyboard.get_key_date(id_tranz, d_new.day, d_new.month, d_new.year)
    )


# Обработка Коментарий trancaction
@form_router.callback_query(keyboards.keyboard.Itog_Callback.filter(F.action == "tr_caments"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.Itog_Callback,
        state: FSMContext
) -> None:
    id_tranz = callback_data.id_tranz
    user = callback_data.val
    data = await sqlite.sel_tranc(id_tranz, user)

    cash = ""
    cat = ""
    date = ""
    for i in data:
        cash = i[0]
        cat = i[1]
        date = i[2]

    await callback.message.edit_text(
        f"💬 Введите, пожалуйста, комментарий."
    )
    await state.set_state(Form.coments)

    await state.update_data(cat=cat)
    await state.update_data(us_id=user)
    await state.update_data(cash=cash)
    await state.update_data(dates=date)
    await state.update_data(id_tr=id_tranz)


##################################################
############ Коментарий

@form_router.message(Form.coments)
# async def process_name(callback: types.CallbackQuery, state: FSMContext) -> None:
async def process_name(message: Message, state: FSMContext) -> None:
    coments = message.text

    data = await state.get_data()
    cash = data["cash"]
    cat = data["cat"]
    us_id = data["us_id"]
    dates = data["dates"]

    idi = data["id_tr"]

    await sqlite.update_trans_com(coments, idi)

    await message.answer(
        text=f"👌 Комментарий успешно добавлен.\n Добавлено {cash} ₽ в категорию {cat} \n  операция расход \n {dates}, <i><b>{coments}</b></i>",
        reply_markup=keyboards.keyboard.get_keyboard(idi, message.message_id, us_id)
        )
    await state.clear()
    print(f"ID Mess  coments {coments}")


# Обработка del trancaction
@form_router.callback_query(keyboards.keyboard.Itog_Callback.filter(F.action == "inc_tr_kl"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.Itog_Callback,
        state: FSMContext
) -> None:
    data = await state.get_data()
    cash = data["cash"]
    sell = data["language"]
    cat = data["name"]
    us_id = data["id_us"]
    flight = data["id_flight"]
    name_flight = data["name_flight"]

    znac = callback_data.val
    mes_id = data["mes_id"]

    # print(f"Del_tranc {znac}")
    if znac == 1:
        print(f"date edit {data}")


async def select_cat(message: types.Message, cat: str, sellect: str, cash: str, us_id: int, edit: int, flight: int,
                     fuel: int):
    with suppress(TelegramBadRequest):

        if edit == 0:
            dann = await sqlite.db_insert_transactions(cat, us_id, cash, now1, flight, fuel)
            idi = 0
            users = 0
            for i in dann:
                idi = i[0]

            await message.edit_text(
                f"Добавлено {cash} ₽ в категорию {cat} операция расход \n {transform_date(now1)} ",
                reply_markup=keyboards.keyboard.get_keyboard(idi, message.message_id, us_id)
            )
            await message.edit_text(
                f"Добавлено {cash} ₽ в категорию {cat} операция расход \n {transform_date(now1)} ",
                reply_markup=keyboards.keyboard.get_keyboard(idi, message.message_id, us_id)
            )
            # print(f"id {id}")
            # print(f"message_id {message.message_id}")


# @form_router.message(CommandStart())
# async def command_start(message: Message, state: FSMContext) -> None:
#     id = message.from_user.id
#     data = await sqlite.db_sel_driver(id)
#     user = ""
#     if len(data) != 0:
#         for i in data:
#             user = i[0]
#         await message.answer(f"Приветствую <b>{user}</b>!\nВы зарегистрированы в системе, можете продолжать работать.")
#     else:
#         await state.set_state(Form.name)
#         await message.answer(
#             "Добро пожаловать!\nПредставьтесь полным ФИО, будем использовать для отчёта.\n"
#             "   Этот бот может помочь водителю в учете доходов и расходов, а также создании отчетов о рейсах.",
#             reply_markup=ReplyKeyboardRemove(),
#         )


@form_router.message(Form.name)
async def handle_code(message: types.Message, state: FSMContext) -> None:
    user_name = message.text
    user_id = message.from_user.id
    await sqlite.db_insert_driver(user_id, user_name)
    await state.clear()
    await message.answer(
        f"Поздравляю <b>{user_name}</b>!\nВы зарегистрированы в системе, можете продолжать работать.",
        reply_markup=ReplyKeyboardRemove(),
    )
    # await message.reply("Я тебя не понимаю :(")


# ОБРАБОТКА К Приходу
@form_router.callback_query(keyboards.keyboard.NumbersCallbackFactory.filter(F.action == "incr"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.NumbersCallbackFactory,
        state: FSMContext
) -> None:
    data = await state.get_data()
    cash = data["cash"]
    name_flight = data["name_flight"]
    users_id = callback.message.chat.id
    # Запрос категории

    data = await sqlite.db_sel_cat(users_id, "incr")

    await callback.message.edit_text(
        f"{name_flight}\nВыбери куда записать: {cash} ₽",
        reply_markup=keyboards.keyboard.gen_markup_deb(data)
    )


# ОБРАБОТКА К Расходу
@form_router.callback_query(keyboards.keyboard.NumbersCallbackFactory.filter(F.action == "decr"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.NumbersCallbackFactory,
        state: FSMContext
) -> None:
    data = await state.get_data()
    cash = data["cash"]
    users_id = callback.message.chat.id

    # Запрос категории
    data = await sqlite.db_sel_cat(users_id, "decr")

    await callback.message.edit_text(
        f"Выбери куда записать: {cash} ₽",
        reply_markup=keyboards.keyboard.gen_markup(data)
    )


# Запрос списка категорий для редактирование
@form_router.message(Command("categories"))
async def buy(message: types.Message):
    data = await sqlite.db_sel_edit_cat(message.chat.id)

    if len(data) != 0:
        await message.answer(f"Список категорий", reply_markup=keyboards.keyboard.gen_list_cat(data))
    else:
        await message.answer(
            f"В вашем списке категорий не обнаружено, предлагаю создать свои или заполнить их по умолчанию.",
            reply_markup=keyboards.keyboard.gen_list_cat_nul()
        )


# Заполняем категории по умолчанию
@form_router.callback_query(keyboards.keyboard.NumbersCallbackFactory.filter(F.action == "ins_cat_def"))
async def command_start(callback: types.CallbackQuery) -> None:
    dan = await sqlite.db_sel_default_cat()
    print(dan)
    print(callback.message.chat.id)
    cat = ""
    wr = ""
    for i in dan:
        cat = i[0]
        wr = i[1]
        await sqlite.db_insert_cat(cat, wr, callback.message.chat.id, i[2])

    await callback.message.edit_text(
        "Категории заполнены"
    )


# Переход категорий для редактирование
@form_router.callback_query(keyboards.keyboard.NumbersCallbackFactory.filter(F.action == "edit_cats"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.NumbersCallbackFactory
) -> None:
    name_cat = callback_data.name
    wiring = callback_data.val
    id_cat = callback_data.value
    wiring_e = ""
    wiring_n = ""
    if wiring == "decr":
        wiring_e = "(расход)"
        wiring_n = "приход"
    elif wiring == "incr":
        wiring_e = "(приход)"
        wiring_n = "расход"

    await callback.message.edit_text(
        f"Редактирование категории <b>{name_cat} {wiring_e}</b>",
        reply_markup=keyboards.keyboard.gen_edits_cat(id_cat, wiring_n, name_cat, wiring)
    )


# Запрос списка категорий для редактирование EXIT
@form_router.callback_query(keyboards.keyboard.NumbersCallbackFactory.filter(F.action == "edit_cats_ex"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
) -> None:
    await callback.message.edit_text("Список категорий\n---\n Выход")


# Удоляем категорию
@form_router.callback_query(keyboards.keyboard.NumbersCallbackFactory.filter(F.action == "edit_cats_del"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.NumbersCallbackFactory
) -> None:
    name_cat = callback_data.name
    id = callback_data.value
    await sqlite.db_del_edit_cat(id)
    await callback.message.edit_text(f'Категория "{name_cat}" \n---\n❌ Удалена')


# Назначаем категорию как топливную
@form_router.callback_query(keyboards.keyboard.NumbersCallbackFactory.filter(F.action == "edit_cats_fuel"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.NumbersCallbackFactory
) -> None:
    name_cat = callback_data.name
    id = callback_data.value
    await sqlite.db_upg_edit_cat(id, "yes", "decr")
    await callback.message.edit_text(f"Назначили {name_cat} как топливную категорию  \n---\n")


# Изменяем в категории операцию
@form_router.callback_query(keyboards.keyboard.NumbersCallbackFactory.filter(F.action == "edit_cats_wiring"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.NumbersCallbackFactory
) -> None:
    name_cat = callback_data.name
    id = callback_data.value
    wiring = callback_data.val
    wiring_e = ""
    wiring_n = ""
    if wiring == "decr":
        wiring_e = "incr"
        wiring_n = "приходную"
    elif wiring == "incr":
        wiring_e = "decr"
        wiring_n = "расходную"

    await sqlite.db_upg_edit_cat_wr(id, wiring_e)
    await callback.message.edit_text(f'Назначили категорию: "{name_cat}" как {wiring_n}  \n---\n')


@form_router.message(Command("cattegor_incr"))
async def buy(message: types.Message):
    data = await sqlite.db_sel_inc()
    await message.answer(f"Список: {data}", reply_markup=keyboards.keyboard.gen_markup(data))


@form_router.callback_query(keyboards.keyboard.NumbersCallbackFactory.filter(F.action == "ins_cat"))
async def command_start(callback: types.CallbackQuery, state: FSMContext) -> None:
    await state.set_state(Form.cat_new)
    await callback.message.edit_text(
        "Введите название категории"
    )


@form_router.message(Form.cat_new)
async def process_name(message: Message, state: FSMContext) -> None:
    await state.update_data(cat_new=message.text)
    await state.update_data(id_us=message.from_user.id)

    await message.answer(
        f"Как будем учитывать категорию, {html.quote(message.text)} ?",
        reply_markup=get_and_cat()
    )


def get_and_cat():
    buttons = [
        [
            types.InlineKeyboardButton(text="Расход", callback_data="num_decr"),
            types.InlineKeyboardButton(text="Приход", callback_data="num_incr")
        ],

    ]
    keyboard = types.InlineKeyboardMarkup(inline_keyboard=buttons)
    return keyboard


@form_router.message(Command("flight"))
async def process_name(message: Message, state: FSMContext) -> None:
    dann = await sqlite.db_sel_flight("active", message.from_user.id)
    dan_repair = await sqlite.db_sel_flight("repair", message.from_user.id)
    rep = ""
    if len(dan_repair) != 0:
        rep = dan_repair

    # print(f"dan_repair: {dan_repair} rep: {rep}")

    if not dann:
        dat = "🛣 Открыть"
        act = "fli_new"
        await message.answer(
            f"<b>Управление рейсами.</b> \nУ вас нет активных рейсов.",
            reply_markup=keyboards.keyboard.key_fl_config(dat, act, rep)
        )

    else:
        dat = "🛣 Завершить"
        act = "fli_off"
        for i in dann:
            await message.answer(
                f"<b>Управление рейсами.</b> \nЗавершить рейс: <b>{i[1]} от {i[2]} </b>",
                reply_markup=keyboards.keyboard.key_fl_config_activ(dat, act, i[0], rep)

            )
        print(dann)
        await state.update_data(user_id_fli=message.from_user.id)


@form_router.callback_query(keyboards.keyboard.Flight_Callback.filter(F.action == "order_repair"))
async def cancel_handler(callback: types.CallbackQuery,
                         callback_data: keyboards.keyboard.Flight_Callback,
                         state: FSMContext) -> None:
    user_id = callback.message.chat.id
    flight_id = callback_data.id_fl
    status_fl = callback_data.st_fl

    await sqlite.db_repair_flight(flight_id, "repair")
    await callback.message.edit_text("Рейс отложен, авто отправили в ремонт")


@form_router.callback_query(keyboards.keyboard.Flight_Callback.filter(F.action == "repair_off"))
async def cancel_handler(callback: types.CallbackQuery,
                         callback_data: keyboards.keyboard.Flight_Callback,
                         state: FSMContext) -> None:
    user_id = callback.message.chat.id
    flight_id = callback_data.id_fl
    name_fl = callback_data.st_fl

    await sqlite.db_repair_flight(flight_id, "active")
    await callback.message.edit_text(f"Рейс {name_fl}, назначен активным.")


@form_router.message(Command("list_cat"))
async def process_name(message: Message, state: FSMContext) -> None:
    dann = await sqlite.db_sel_flight("active", message.from_user.id)
    await message.answer(
        "List categor",
        reply_markup=keyboards.keyboard.gen_list_cat()
    )


############ Запрос списка рейсов
@form_router.callback_query(keyboards.keyboard.Flight_Callback.filter(F.action == "list_flights"))
async def callbacks_num_change_fab(callback: types.CallbackQuery) -> None:
    users = callback.message.chat.id
    data = await sqlite.db_sel_flight_name(users)

    await callback.message.edit_text(
        f"<b>Список последних 10 рейсов</b>",
        reply_markup=keyboards.keyboard.gen_list_flights(data)
    )


############ Запрос списка поступлений
@form_router.callback_query(keyboards.keyboard.Flight_Callback.filter(F.action == "list_receipt"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.Flight_Callback
) -> None:
    users = callback.message.chat.id
    id_fl = callback_data.id_fl
    data = await sqlite.db_sel_tr_pay(id_fl, "incr")

    await callback.message.edit_text(
        f"<b>Список поступлений в рейсе</b>",
        reply_markup=keyboards.keyboard.gen_list_pay(data)
    )


############ Запрос списка оплат
@form_router.callback_query(keyboards.keyboard.Flight_Callback.filter(F.action == "list_pay"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.Flight_Callback
) -> None:
    users = callback.message.chat.id
    id_fl = callback_data.id_fl

    data = await sqlite.db_sel_tr_pay(id_fl, "decr")

    await callback.message.edit_text(
        f"<b>Список оплат в рейсе</b>",
        reply_markup=keyboards.keyboard.gen_list_pay(data)
    )


############ Запрос списка заправок
@form_router.callback_query(keyboards.keyboard.Flight_Callback.filter(F.action == "list_fuels"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.Flight_Callback
) -> None:
    users = callback.message.chat.id
    id_fl = callback_data.id_fl
    data = await sqlite.db_sel_tr_fuels(id_fl)
    print(data)
    litr = 0
    if len(data) != 0:
        for i in data:
            litr = litr + i[2]

        await callback.message.edit_text(
            f"<b>⛽ Список заправок в рейсе</b>\nЗаправлено: <b>{litr}л.</b>",
            reply_markup=keyboards.keyboard.gen_list_fuels(data)
        )
    else:
        await callback.message.edit_text(f"<b>⛽ Учтённых заправок нет.</b>")


############ В Рейс
@form_router.callback_query(keyboards.keyboard.Flight_Callback.filter(F.action == "fli_new"))
async def callbacks_num_change_fab(callback: types.CallbackQuery) -> None:
    data = await sqlite.db_sel_cars()
    await callback.message.edit_text(
        "Выбери транспорт",
        reply_markup=keyboards.keyboard.gen_select_cars(data)
    )


@form_router.callback_query(keyboards.keyboard.NumbersCallbackFactory.filter(F.action == "selcars"))
async def process_name_flight(callback: types.CallbackQuery,
                              callback_data: keyboards.keyboard.NumbersCallbackFactory,
                              state: FSMContext) -> None:
    users = callback.message.chat.id
    print(f"ID users: {users}")

    await state.update_data(carsM=callback_data.val)
    await state.update_data(cars=callback_data.value)
    data = await sqlite.db_sel_nam_point(users)
    await callback.message.edit_text(f"<b>Куда едим ?</b>\nВыбери из списка или введи новый пункт.",
                                     reply_markup=keyboards.keyboard.gen_select_point(data))
    await state.set_state(Flight.name)


############ Обработка Выбора Конечной точкт Рейса
@form_router.callback_query(keyboards.keyboard.Callback_Point.filter(F.action == "point"))
async def process_name_flight(callback: types.CallbackQuery,
                              callback_data: keyboards.keyboard.Callback_Point,
                              state: FSMContext) -> None:
    await state.update_data(name=callback_data.value)
    await callback.message.answer(
        text=f"Рейс : {callback_data.value} \nВведи начальный километраж.",
        reply_markup=ReplyKeyboardRemove(),
    )
    await state.set_state(Flight.mileage)


############ Обработка ввода Конечной точкт Рейса
@form_router.message(Flight.name)
async def process_name_flight(message: Message, state: FSMContext) -> None:
    await state.update_data(name=message.text)
    name = message.text
    await message.answer(
        text=f"Рейс : {name} \nВведи начальный километраж.",
        reply_markup=ReplyKeyboardRemove(),
    )
    await state.set_state(Flight.mileage)


############ Обработка ввода начального пробега
@form_router.message(Flight.mileage)
async def process_write_mileage(message: Message, state: FSMContext) -> None:
    await state.update_data(mileage=message.text)
    await message.answer(
        text=f"Введи остаток топлива.",
        reply_markup=ReplyKeyboardRemove(),
    )
    await state.set_state(Flight.fuel)


############ Обработка ввода начального остатка топлива
@form_router.message(Flight.fuel)
async def process_write_mileage(message: Message, state: FSMContext) -> None:
    await state.update_data(fuel=message.text)
    await state.update_data(id_us=message.from_user.id)
    data = await state.get_data()
    mileage = data["mileage"]
    name = data["name"]
    fuel_in = data["fuel"]
    cars = data["cars"]
    carsMarka = data["carsM"]
    await message.answer(
        text=f"Рейс : {name} от {transform_date(now1)}\nТранспорт: <b>{carsMarka}</b>\nНачальный километраж: <b>{mileage}</b>\n Остаток топлива: <b>{fuel_in}</b>\n Верно?",
        reply_markup=keyboards.keyboard.key_flight(),
    )


@form_router.callback_query(keyboards.keyboard.Flight_Callback.filter(F.action == "fli_okdist"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        state: FSMContext
) -> None:
    data = await state.get_data()
    mileage = data["mileage"]
    name = data["name"]
    fuel = data["fuel"]
    id_us = data["id_us"]
    cars = data["cars"]
    carsMarka = data["carsM"]
    status = "active"
    await callback.message.edit_text(
        f"записал:\n Рейс : {name} от {transform_date(now1)}\n Транспорт: <b>{carsMarka}</b>\nНачальный километраж: <b>{mileage}</b>\n Остаток топлива: {fuel}."
    )
    await sqlite.db_insert_flight(name, id_us, fuel, mileage, now, cars, status)
    await state.clear()


@form_router.callback_query(keyboards.keyboard.Flight_Callback.filter(F.action == "now_cars"))
async def process_name_flight(callback: types.CallbackQuery, state: FSMContext) -> None:
    await callback.message.edit_text("Отмена")
    await state.clear()


############/ В Рейс /##########################################################

############ Из Рейса /##########################################################

@form_router.callback_query(keyboards.keyboard.Flight_Callback.filter(F.action == "fli_off"))
async def process_name(callback: types.CallbackQuery, state: FSMContext) -> None:
    dat = await state.get_data()
    user_id_fli = dat["user_id_fli"]

    dann = await sqlite.db_sel_flight("active", user_id_fli)

    for i in dann:
        await callback.message.edit_text(
            f"Завершить рейс: <b>{i[1]} от {i[2]} </b>",
            reply_markup=get_key_fl()
        )

        await state.update_data(name=i[1] + " от " + i[2])
        await state.update_data(id=i[0])
        await state.update_data(date_in=i[2])
        await state.update_data(in_fuel=i[4])
        await state.update_data(in_odometr=i[5])
        await state.update_data(cars=i[3])


def get_key_fl():
    buttons = [
        [
            types.InlineKeyboardButton(text="Да", callback_data="exit_flight"),
            types.InlineKeyboardButton(text="Нет", callback_data="num_incr")
        ],
    ]
    keyboard = types.InlineKeyboardMarkup(inline_keyboard=buttons)
    return keyboard


@form_router.callback_query(F.data.startswith("exit_flight"))
async def callbacks_num(callback: types.CallbackQuery, state: FSMContext):
    await callback.message.edit_text("Введи конечный километраж.")
    await state.set_state(Flight_exit.finish_odometr)


@form_router.message(Flight_exit.finish_odometr)
async def process_write_mileage(message: Message, state: FSMContext) -> None:
    print(message.text)
    await state.update_data(odometr=message.text)
    await message.answer(
        text=f"Введи остаток топлива.",
        reply_markup=ReplyKeyboardRemove(),
    )
    await state.set_state(Flight_exit.fuel_residues)


@form_router.message(Flight_exit.fuel_residues)
async def process_write_mileage(message: Message, state: FSMContext) -> None:
    ful = await state.update_data(fuel=message.text)
    dat = await state.get_data()
    id = dat["id"]
    carsMarka = dat["cars"]
    name = dat["name"]
    odometr = dat["odometr"]
    fuel = dat["fuel"]
    status = "fulfilled"
    #  await sqlite.db_obdate_flight(id,fuel,odometr,now,status)

    await message.answer(
        f"Закрываем рейс с данными?\n Рейс : {name} \n Транспорт: {carsMarka}\nКонечный километраж: {odometr}\n Остаток топлива: {fuel}.",
        reply_markup=keyboards.keyboard.key_flight_confirmation()
    )


@form_router.callback_query(keyboards.keyboard.Flight_Callback.filter(F.action == "fli_confirmation"))
async def process_like_write_bots(callback: types.CallbackQuery,
                                  callback_data: keyboards.keyboard.NumbersCallbackFactory,
                                  state: FSMContext) -> None:
    dat = await state.get_data()
    id = dat["id"]
    carsMarka = dat["cars"]
    name = dat["name"]

    in_odometr = dat["in_odometr"]
    odometr = dat["odometr"]
    Mileage_covered = int(odometr) - int(in_odometr)

    in_fuel = dat["in_fuel"]
    fuel = dat["fuel"]
    Fuel_used_up = int(in_fuel) - int(fuel)

    Fuel_consumption = round(Fuel_used_up / Mileage_covered * 100, 2)

    status = "fulfilled"
    await sqlite.db_obdate_flight(id, fuel, odometr, now, status)
    print(id)
    await exel_creat(id)

    await callback.message.edit_text(
        f"Записал. \nРейс : {name} Пройденный километраж: {Mileage_covered}\n Израсходовано топлива: {Fuel_used_up}\n Расход топлива: {Fuel_consumption}"
    )
    # await exel_creat(id)
    await state.clear()


async def update_num_text(message: types.Message, new_value: str):
    await message.edit_text(
        f"Уитывать категорию: {new_value} как",
        reply_markup=get_keyboard()
    )


def get_keyboard():
    buttons = [
        [
            types.InlineKeyboardButton(text="Расход", callback_data="num_decr"),
            types.InlineKeyboardButton(text="Приход", callback_data="num_incr")
        ],
        [types.InlineKeyboardButton(text="Подтвердить", callback_data="num_finish")]
    ]
    keyboard = types.InlineKeyboardMarkup(inline_keyboard=buttons)
    return keyboard


###################################################
@form_router.callback_query(F.data.startswith("num_"))
async def callbacks_num(callback: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    us_id = data["id_us"]
    cat_new = data["cat_new"]
    action = callback.data.split("_")[1]
    await state.clear()

    if action == "incr":
        await sqlite.db_insert_cat(cat_new, "incr", us_id, " ")
        await callback.message.edit_text(f'Добавлена категория : <b>"{cat_new}"</b> как поступление.')

    elif action == "decr":
        await sqlite.db_insert_cat(cat_new, "decr", us_id, "")
        await callback.message.edit_text(f'Добавлена категория : <b>"{cat_new}"</b> как расход.')

    elif action == "finish":
        await sqlite.db_insert_cat(cat_new, "finish", us_id)
        await callback.message.edit_text(f"Итого: {us_id} finish {cat_new}")

    await callback.answer()


@form_router.message(Form.cat_new, F.text.casefold() == "No")
async def process_like_write_bots(message: Message, state: FSMContext) -> None:
    await state.clear()
    await message.reply(
        "Круто! Я тоже в Расход",
        reply_markup=ReplyKeyboardRemove(),
    )


# @form_router.message(Command("cancel"))
@form_router.message(F.text.casefold() == "cancel")
async def cancel_handler(message: Message, state: FSMContext) -> None:
    """
    Allow user to cancel any action
    """
    current_state = await state.get_state()
    if current_state is None:
        return

    logging.info("Cancelling state %r", current_state)
    await state.clear()
    await message.answer(
        "Cancelled.",
        reply_markup=ReplyKeyboardRemove(),
    )


###### ОТЧЁТ
# @form_router.message(Command("order"))
@form_router.callback_query(keyboards.keyboard.Flight_Callback.filter(F.action == "order_flights"))
async def cancel_handler(callback: types.CallbackQuery,
                         callback_data: keyboards.keyboard.Flight_Callback,
                         state: FSMContext) -> None:
    user_id = callback.message.chat.id
    flight_id = callback_data.id_fl
    status_fl = callback_data.st_fl

    fuel = 0
    fuels = 0
    # print(f"id {user_id} fl_id: {flight_id}")
    order_str = ""
    sum_incr = 0
    sum_decr = 0
    driver = ""
    order_str = order_str + "<b>Поступление:</b>\n"
    # Запросы
    incrdat = await sqlite.db_sel_tr_order(flight_id, "incr")
    decrdat = await sqlite.db_sel_tr_order(flight_id, "decr")
    odometr = await sqlite.db_order_info(flight_id)
    drivers = await sqlite.db_sel_driver(user_id)
    # print(f"drivers {drivers}")
    for i in drivers:
        driver = i[0]

    for i in incrdat:
        order_str = order_str + f"  {i[1]} Σ: {i[2]}р.\n"
        sum_incr = sum_incr + i[2]
    order_str = order_str + f"<b>Итого</b> получено: <b>{sum_incr}р.</b>\n"

    order_str = order_str + "\n<b>Расходы:</b>\n"
    for i in decrdat:
        order_str = order_str + f"  {i[1]} Σ: {i[2]}р.\n"
        sum_decr = sum_decr + i[2]
        if i[3] != 0.0:
            fuel = i[3]

    order_str = order_str + f"<b>Итого</b> расходов: <b>{sum_decr}р.</b>\n"
    order_str = order_str + f"<b>Остатк</b> ДС: <b>{sum_incr - sum_decr}р.</b>\n"
    order_str = order_str + f"<b>\nТопливо:</b>\n"
    # odom_in = 0
    # odom_ex = 0
    odom = 0
    auto = ""
    for i in odometr:
        order_str = order_str + f"  Начальный остаток: {i[0]} л.\n  Заправлено: {fuel} л.\n"
        fuels = fuel + i[0]

        if i[1] != None:
            fuels = fuels - i[1]
            order_str = order_str + f"  Конечный остаток: {i[1]} л.\n<b>Израсходовано:</b> {fuels} л.\n\n<b>Одометр\n</b>"

        else:
            order_str = order_str + f"\n\n<b>Одометр\n</b>"

        order_str = order_str + f"  Начальный: {i[2]} km.\n"
        odom = i[2]

        if i[3] != None:
            # odom_ex = i[3]
            # odom = odom_ex - odom_in
            odom = abs(i[3] - odom)
            order_str = order_str + f"  Конечный: {i[3]} km.\n<b>Пройдено : {odom} km.</b>"
            order_str = order_str + f"\n \n<b>Средний расход:</b> {round(fuels / odom * 100, 2)} л/100км"

        auto = i[6]
        order_str = order_str + f"\nАвто: {auto} {i[7]}"
        order_str = order_str + f"\nВодитель: {driver}"

    if status_fl == "fulfilled":
        await callback.message.edit_text(order_str,
                                         reply_markup=keyboards.keyboard.gen_key_exel(flight_id)
                                         )

    else:
        await callback.message.edit_text(order_str)


@form_router.message(Form.like_bots)
async def process_unknown_write_bots(message: Message) -> None:
    await message.reply("Я тебя не понимаю :(")


@form_router.message(F.text.regexp(r"(\d+)", mode=RegexpMode.SEARCH).as_("code"))
async def handle_code(message: types.Message, code: Match[str], state: FSMContext) -> None:
    await state.update_data(id_us=message.from_user.id)

    dann = await sqlite.db_sel_flight("active", message.from_user.id)
    if not dann:
        await message.answer(
            f"У вас нет активных рейсов!",
            reply_markup=keyboards.keyboard.key_fl_new()
        )
    else:

        for i in dann:
            await state.update_data(name_flight=f"Рейс: <b>{i[1]}</b> от {i[2]} ")
            await state.update_data(id_flight=i[0])

        name = await state.get_data()
        name_flight = name["name_flight"]
        data = await state.update_data(cash=code.group())
        cash = data["cash"]
        await state.set_state(Form.cash)
        data_cat = await sqlite.db_sel_cat(message.from_user.id, "decr")
        if len(data_cat) != 0:
            await message.answer(
                f"{name_flight}\nВыбери куда записать: <b>{cash}</b> ₽",
                reply_markup=keyboards.keyboard.gen_markup(data_cat)
            )
        else:
            await message.answer(
                f"Категорий учета отсутствуют. Предлагаю создать свои каткгории или заполнить списоком по умолчанию, а затем повторить ввод суммы.",
                reply_markup=keyboards.keyboard.gen_list_cat_nul()
            )

    data = await state.get_data()
    cash = data["cash"]
    name_flight = data["name_flight"]
    # Запрос категории
    data = await sqlite.db_sel_cat(message.from_user.id, "incr")


async def main():
    bot = Bot(token=TOKEN, parse_mode=ParseMode.HTML)
    dp = Dispatcher()
    dp.include_router(form_router)

    await dp.start_polling(bot,
                           on_startup=on_startup
                           )


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO,
                        stream=sys.stdout,
                        )
    asyncio.run(main())

from aiogram.fsm.state import State, StatesGroup


class Registration(StatesGroup):
    choosing_role = State()
    entering_name = State()


@form_router.message(CommandStart())

async def start_handler(message: Message, state: FSMContext):
    user_id = message.from_user.id
    print("us")
    conn = sqlite3.connect("my_database.db")
    cursor = conn.cursor()
    print("ssfdsd")
    cursor.execute("SELECT role FROM users WHERE tg_id = ?", (user_id,))
    user = cursor.fetchone()

    if user:
        await message.answer("Вы уже зарегистрированы.")
    else:
        keyboard = ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="🚛 Водитель")],
                [KeyboardButton(text="🏢 Владелец парка")]
            ],
            resize_keyboard=True
        )
        await message.answer("Привет! Выберите вашу роль:", reply_markup=keyboard)
        await state.set_state(Registration.choosing_role)

    conn.close()


@form_router.message(Registration.choosing_role)
async def process_role_choice(message: Message, state: FSMContext):
    role = message.text.strip()

    if role not in ["🚛 Водитель", "🏢 Владелец парка"]:
        await message.answer("Выберите роль, используя кнопки ниже.")
        return

    role_db = "driver" if role == "🚛 Водитель" else "owner"
    await state.update_data(role=role_db)

    await message.answer("Теперь введите ваше имя:")
    await state.set_state(Registration.entering_name)
