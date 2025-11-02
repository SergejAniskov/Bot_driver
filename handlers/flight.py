
import sqlite
import keyboards.keyboard
from aiogram import types
from aiogram import Bot, Dispatcher, F, Router, html
from aiogram.filters import Command, CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from datetime import datetime, timedelta, date
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
import pandas as pd
import emoji
import subprocess

from openpyxl.worksheet import worksheet
from win32com import client
from aiogram.enums import ParseMode, ChatAction


from aiogram.types import (
    KeyboardButton,
    Message,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
)

# from main import transform_date, now1, now, exel_creat


class orders_gen(StatesGroup):
    namber = State()


class Flight(StatesGroup):
    name = State()
    status = State()
    cars = State()
    mileage = State()
    fuel = State()

class Flight_exit(StatesGroup):
    status = State()
    name = State()
    finish_odometr = State()
    fuel_residues = State()

form_router = Router()

@form_router.message(Command("flight"))
async def process_name(message: Message, state: FSMContext) -> None:
    dann = await sqlite.db_sel_flight("active",message.from_user.id)
    dan_repair = await sqlite.db_sel_flight("repair", message.from_user.id)
    rep = ""
    if len(dan_repair) != 0:
        rep = dan_repair

    # print(f"dan_repair: {dan_repair} rep: {rep}")

    if not dann:
        dat = "🛣 Открыть"
        act = "fli_new"
        await message.answer(
            f"<b>Управление рейсами.</b> \nУ вас нет активных рейсов.",
                reply_markup=keyboards.keyboard.key_fl_config(dat,act,rep)
                )

    else:
        dat = "🛣 Завершить"
        act = "fli_off"
        for i in dann:

            await message.answer(
                f"<b>Управление рейсами.</b> \nЗавершить рейс: <b>{i[1]} от {i[2]} </b>",
                reply_markup=keyboards.keyboard.key_fl_config_activ(dat,act, i[0],rep)

            )
        print(dann)
        await state.update_data(user_id_fli=message.from_user.id)



# Отложить рейс


@form_router.callback_query(keyboards.keyboard.Flight_Callback.filter(F.action == "order_repair"))
async def cancel_handler(callback: types.CallbackQuery,
                         callback_data: keyboards.keyboard.Flight_Callback,
                         state: FSMContext) -> None:
    user_id = callback.message.chat.id
    flight_id = callback_data.id_fl
    status_fl = callback_data.st_fl

    await sqlite.db_repair_flight(flight_id,"repair")
    await callback.message.edit_text("Рейс отложен, авто отправили в ремонт")


@form_router.callback_query(keyboards.keyboard.Flight_Callback.filter(F.action == "repair_off"))
async def cancel_handler(callback: types.CallbackQuery,
                         callback_data: keyboards.keyboard.Flight_Callback,
                         state: FSMContext) -> None:
    user_id = callback.message.chat.id
    flight_id = callback_data.id_fl
    name_fl = callback_data.st_fl

    await sqlite.db_repair_flight(flight_id,"active")
    await callback.message.edit_text(f"Рейс {name_fl}, назначен активным.")


############ Запрос списка рейсов
@form_router.callback_query(keyboards.keyboard.Flight_Callback.filter(F.action == "list_flights"))
async def callbacks_num_change_fab(callback: types.CallbackQuery) -> None:
    users = callback.message.chat.id
    data = await sqlite.db_sel_flight_name(users)

    await callback.message.edit_text(
        f"<b>Список последних 10 рейсов</b>",
        reply_markup=keyboards.keyboard.gen_list_flights(data)
    )

############ Запрос списка поступлений
@form_router.callback_query(keyboards.keyboard.Flight_Callback.filter(F.action == "list_receipt"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.Flight_Callback
) -> None:
    users = callback.message.chat.id
    id_fl = callback_data.id_fl
    data = await sqlite.db_sel_tr_pay(id_fl,"incr")

    await callback.message.edit_text(
        f"<b>Список поступлений в рейсе</b>",
        reply_markup=keyboards.keyboard.gen_list_pay(data)
    )

############ Запрос списка оплат
@form_router.callback_query(keyboards.keyboard.Flight_Callback.filter(F.action == "list_pay"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.Flight_Callback
) -> None:
    users = callback.message.chat.id
    id_fl = callback_data.id_fl

    data = await sqlite.db_sel_tr_pay(id_fl,"decr")

    await callback.message.edit_text(
        f"<b>Список оплат в рейсе</b>",
        reply_markup=keyboards.keyboard.gen_list_pay(data)
    )

############ Запрос списка заправок
@form_router.callback_query(keyboards.keyboard.Flight_Callback.filter(F.action == "list_fuels"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        callback_data: keyboards.keyboard.Flight_Callback
        ) -> None:
    users = callback.message.chat.id
    id_fl = callback_data.id_fl
    data = await sqlite.db_sel_tr_fuels(id_fl)
    print(data)
    litr = 0
    if len(data) != 0:
        for i in data:
            litr = litr + i[2]

        await callback.message.edit_text(
            f"<b>⛽ Список заправок в рейсе</b>\nЗаправлено: <b>{litr}л.</b>",
            reply_markup=keyboards.keyboard.gen_list_fuels(data)
        )
    else:
        await callback.message.edit_text(f"<b>⛽ Учтённых заправок нет.</b>" )


############ В Рейс
@form_router.callback_query(keyboards.keyboard.Flight_Callback.filter(F.action == "fli_new"))
async def callbacks_num_change_fab(callback: types.CallbackQuery) -> None:
    data = await sqlite.db_sel_cars()
    await callback.message.edit_text(
        "Выбери транспорт",
        reply_markup=keyboards.keyboard.gen_select_cars(data)
        )



@form_router.callback_query(keyboards.keyboard.NumbersCallbackFactory.filter(F.action == "selcars"))
async def process_name_flight(callback: types.CallbackQuery,
                                  callback_data: keyboards.keyboard.NumbersCallbackFactory,
                                  state: FSMContext) -> None:
    users = callback.message.chat.id
    print(f"ID users: {users}")

    await state.update_data(carsM=callback_data.val)
    await state.update_data(cars=callback_data.value)
    data = await sqlite.db_sel_nam_point(users)
    await callback.message.edit_text(f"<b>Куда едим ?</b>\nВыбери из списка или введи новый пункт.", reply_markup=keyboards.keyboard.gen_select_point(data))
    await state.set_state(Flight.name)


############ Обработка Выбора Конечной точкт Рейса
@form_router.callback_query(keyboards.keyboard.Callback_Point.filter(F.action == "point"))
async def process_name_flight(callback: types.CallbackQuery,
                                  callback_data: keyboards.keyboard.Callback_Point,
                                  state: FSMContext) -> None:

    await state.update_data(name=callback_data.value)
    await callback.message.answer(
        text=f"Рейс : {callback_data.value} \nВведи начальный километраж.",
        reply_markup=ReplyKeyboardRemove(),
    )
    await state.set_state(Flight.mileage)


############ Обработка ввода Конечной точкт Рейса
@form_router.message(Flight.name)
async def process_name_flight(message: Message, state: FSMContext) -> None:
    await state.update_data(name=message.text)
    name = message.text
    await message.answer(
            text=f"Рейс : {name} \nВведи начальный километраж.",
        reply_markup=ReplyKeyboardRemove(),
    )
    await state.set_state(Flight.mileage)

############ Обработка ввода начального пробега
@form_router.message(Flight.mileage)
async def process_write_mileage(message: Message, state: FSMContext) -> None:
    await state.update_data(mileage=message.text)
    await message.answer(
        text=f"Введи остаток топлива.",
        reply_markup=ReplyKeyboardRemove(),
    )
    await state.set_state(Flight.fuel)

############ Обработка ввода начального остатка топлива
@form_router.message(Flight.fuel)
async def process_write_mileage(message: Message, state: FSMContext) -> None:
    await state.update_data(fuel=message.text)
    await state.update_data(id_us=message.from_user.id)
    data = await state.get_data()
    mileage = data["mileage"]
    name = data["name"]
    fuel_in = data["fuel"]
    cars = data["cars"]
    carsMarka = data["carsM"]
    await message.answer(
            text=f"Рейс : {name} от {transform_date(now1)}\nТранспорт: <b>{carsMarka}</b>\nНачальный километраж: <b>{mileage}</b>\n Остаток топлива: <b>{fuel_in}</b>\n Верно?",
        reply_markup=keyboards.keyboard.key_flight(),
    )

now = datetime.now().strftime("%Y-%m-%d %H:%M")
now2 = datetime.now().strftime("%d.%m.%Y")
now1 = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def transform_date(date):
    months = ['января', 'февраля', 'марта', 'апреля', 'мая', 'июня',
              'июля', 'августа', 'сентября', 'октября', 'ноября', 'декабря']
    day, month, year = now2.split('.')
    return f'{day} {months[int(month) - 1]} {year} года'


@form_router.callback_query(keyboards.keyboard.Flight_Callback.filter(F.action == "fli_okdist"))
async def callbacks_num_change_fab(
        callback: types.CallbackQuery,
        state: FSMContext
) -> None:

    data = await state.get_data()
    mileage = data["mileage"]
    name = data["name"]
    fuel = data["fuel"]
    id_us = data["id_us"]
    cars = data["cars"]
    carsMarka = data["carsM"]
    status = "active"
    await callback.message.edit_text(
        f"записал:\n Рейс : {name} от {transform_date(now1)}\n Транспорт: <b>{carsMarka}</b>\nНачальный километраж: <b>{mileage}</b>\n Остаток топлива: {fuel}."
    )
    await sqlite.db_insert_flight(name,id_us,fuel,mileage,now,cars,status)
    await state.clear()



@form_router.callback_query(keyboards.keyboard.Flight_Callback.filter(F.action == "now_cars"))
async def process_name_flight(callback: types.CallbackQuery, state: FSMContext) -> None:

    await callback.message.edit_text("Отмена")
    await state.clear()


############/ В Рейс /##########################################################


############ Из Рейса /##########################################################

@form_router.callback_query(keyboards.keyboard.Flight_Callback.filter(F.action == "fli_off"))
async def process_name(callback: types.CallbackQuery, state: FSMContext) -> None:

    dat = await state.get_data()
    user_id_fli = dat["user_id_fli"]

    dann = await sqlite.db_sel_flight("active", user_id_fli)

    for i in dann:

        await callback.message.edit_text(
            f"Завершить рейс: <b>{i[1]} от {i[2]} </b>",
            reply_markup=get_key_fl()
        )

        await state.update_data(name= i[1]+" от "+ i[2])
        await state.update_data(id = i[0])
        await state.update_data(date_in = i[2])
        await state.update_data(in_fuel=i[4])
        await state.update_data(in_odometr=i[5])
        await state.update_data(cars=i[3])
def get_key_fl():
    buttons = [
        [
            types.InlineKeyboardButton(text="Да", callback_data="exit_flight"),
            types.InlineKeyboardButton(text="Нет", callback_data="num_incr")
        ],
    ]
    keyboard = types.InlineKeyboardMarkup(inline_keyboard=buttons)
    return keyboard

@form_router.callback_query(F.data.startswith("exit_flight"))
async def callbacks_num(callback: types.CallbackQuery, state: FSMContext):
    await callback.message.edit_text("Введи конечный километраж.")
    await state.set_state(Flight_exit.finish_odometr)


@form_router.message(Flight_exit.finish_odometr)
async def process_write_mileage(message: Message, state: FSMContext) -> None:
    print(message.text)
    await state.update_data(odometr=message.text)
    await message.answer(
        text=f"Введи остаток топлива.",
        reply_markup=ReplyKeyboardRemove(),
    )
    await state.set_state(Flight_exit.fuel_residues)

@form_router.message(Flight_exit.fuel_residues)
async def process_write_mileage(message: Message, state: FSMContext) -> None:
    ful = await state.update_data(fuel=message.text)
    dat = await state.get_data()
    id = dat["id"]
    carsMarka = dat["cars"]
    name = dat["name"]
    odometr = dat["odometr"]
    fuel = dat["fuel"]
    status = "fulfilled"
    #  await sqlite.db_obdate_flight(id,fuel,odometr,now,status)

    await message.answer(
        f"Закрываем рейс с данными?\n Рейс : {name} \n Транспорт: {carsMarka}\nКонечный километраж: {odometr}\n Остаток топлива: {fuel}.",
        reply_markup=keyboards.keyboard.key_flight_confirmation()
    )

@form_router.callback_query(keyboards.keyboard.Flight_Callback.filter(F.action == "fli_confirmation"))
async def process_like_write_bots(callback: types.CallbackQuery,
                                  callback_data: keyboards.keyboard.NumbersCallbackFactory,
                                  state: FSMContext) -> None:


    dat = await state.get_data()
    id = dat["id"]
    carsMarka = dat["cars"]
    name = dat["name"]

    in_odometr = dat["in_odometr"]
    odometr = dat["odometr"]
    Mileage_covered = int(odometr) - int(in_odometr)

    in_fuel = dat["in_fuel"]
    fuel = dat["fuel"]
    Fuel_used_up = int(in_fuel)-int(fuel)

    Fuel_consumption = round(Fuel_used_up / Mileage_covered * 100, 2)

    status = "fulfilled"
    await sqlite.db_obdate_flight(id,fuel,odometr,now,status)
    print(id)
    await exel_creat(id)

    await callback.message.edit_text(
        f"Записал. \nРейс : {name} Пройденный километраж: {Mileage_covered}\n Израсходовано топлива: {Fuel_used_up}\n Расход топлива: {Fuel_consumption}"
    )
    # await exel_creat(id)
    await state.clear()


###### ОТЧЁТ
# @form_router.message(Command("order"))
@form_router.callback_query(keyboards.keyboard.Flight_Callback.filter(F.action == "order_flights"))
async def cancel_handler(callback: types.CallbackQuery,
                         callback_data: keyboards.keyboard.Flight_Callback,
                         state: FSMContext) -> None:
    user_id = callback.message.chat.id
    flight_id = callback_data.id_fl
    status_fl = callback_data.st_fl

    fuel = 0
    fuels = 0
    # print(f"id {user_id} fl_id: {flight_id}")
    order_str = ""
    sum_incr = 0
    sum_decr = 0
    driver = ""
    order_str = order_str + "<b>Поступление:</b>\n"
   # Запросы
    incrdat = await sqlite.db_sel_tr_order(flight_id,"incr")
    decrdat = await sqlite.db_sel_tr_order(flight_id, "decr")
    odometr = await sqlite.db_order_info(flight_id)
    drivers = await sqlite.db_sel_driver(user_id)
    # print(f"drivers {drivers}")
    for i in drivers:
        driver = i[0]

    for i in incrdat:
        order_str = order_str + f"  {i[1]} Σ: {i[2]}р.\n"
        sum_incr = sum_incr + i[2]
    order_str = order_str + f"<b>Итого</b> получено: <b>{sum_incr}р.</b>\n"

    order_str = order_str + "\n<b>Расходы:</b>\n"
    for i in decrdat:
        order_str = order_str + f"  {i[1]} Σ: {i[2]}р.\n"
        sum_decr = sum_decr + i[2]
        if i[3] != 0.0:
            fuel = i[3]

    order_str = order_str + f"<b>Итого</b> расходов: <b>{sum_decr}р.</b>\n"
    order_str = order_str + f"<b>Остатк</b> ДС: <b>{sum_incr-sum_decr}р.</b>\n"
    order_str = order_str + f"<b>\nТопливо:</b>\n"
    # odom_in = 0
    # odom_ex = 0
    odom = 0
    auto =""
    for i in odometr:
        order_str = order_str + f"  Начальный остаток: {i[0]} л.\n  Заправлено: {fuel} л.\n"
        fuels = fuel + i[0]

        if i[1] != None:
             fuels = fuels - i[1]
             order_str = order_str + f"  Конечный остаток: {i[1]} л.\n<b>Израсходовано:</b> {fuels} л.\n\n<b>Одометр\n</b>"

        else:
            order_str = order_str + f"\n\n<b>Одометр\n</b>"

        order_str = order_str + f"  Начальный: {i[2]} km.\n"
        odom = i[2]

        if i[3] != None:
            # odom_ex = i[3]
            # odom = odom_ex - odom_in
            odom = abs(i[3] -odom)
            order_str = order_str + f"  Конечный: {i[3]} km.\n<b>Пройдено : {odom} km.</b>"
            order_str = order_str + f"\n \n<b>Средний расход:</b> {round(fuels / odom * 100,2)} л/100км"

        auto = i[6]
        order_str = order_str + f"\nАвто: {auto} {i[7]}"
        order_str = order_str + f"\nВодитель: {driver}"


    if status_fl =="fulfilled":
        await callback.message.edit_text(order_str,
            reply_markup=keyboards.keyboard.gen_key_exel(flight_id)
        )

    else:
        await callback.message.edit_text(order_str)


# Обработка Эксель
async def exel_creat(id_fl: int):

    filename = "templete.xlsx"
    book = openpyxl.load_workbook(filename=filename)
    sheet: worksheet = book.worksheets[0]

    incr_db = await sqlite.db_sel_other_shapka(id_fl)
    auto = ""
    name = ""
    for i in incr_db:
        sheet["B1"].value = i[0]
        name = i[1]
        auto = i[1]

        sheet["B3"].value = auto +" "+ i[2]
        sheet["B4"].value = i[3]

        sheet["J2"].value = i[4]
        date_in = datetime.strptime((i[4]), '%Y-%m-%d %H:%M')
        date_ex = datetime.strptime((i[5]), '%Y-%m-%d %H:%M')
        sheet["L2"].value = i[5]
        s = i[5]

        # print(s[:-6])
        name = name + "От" + s[:-6]
        sheet["J3"].value = i[6]
        sheet["L3"].value = i[7]
        sheet["J4"].value = i[8]
        sheet["L4"].value = i[9]
        sheet["E4"].value = i[10]
        delta_days = (date_ex - date_in).days
        sheet["N2"].value = delta_days



    incr_db = await sqlite.db_sel_tr_pay(id_fl,"incr")
    fuel_db = await sqlite.db_sel_tr_fuels(id_fl)
    other_db = await sqlite.db_sel_tr_other(id_fl)



# # Заполняем поступление
#     print(incr_db)
    nomer = 7
    for i in incr_db:
        pay = i[1]

        if 'карт' in pay:
            sheet.cell(row=nomer, column=2).value = i[2]
        else:
            sheet.cell(row=nomer, column=3).value = i[2]

        date = datetime.strptime((i[3]), '%Y-%m-%d %H:%M:%S')
        sheet.cell(row=nomer, column=1).value = date.strftime('%d-%m-%y %H:%M')
        nomer = nomer + 1


# Заполняем топливо
    nomer = 7
    for i in fuel_db:
        date = datetime.strptime((i[1]), '%Y-%m-%d %H:%M:%S')
        sheet.cell(row=nomer, column=6).value = date.strftime('%d-%m-%y %H:%M')
        sheet.cell(row=nomer, column=7).value = i[2]
        sheet.cell(row=nomer, column=8).value = i[3]
        nomer = nomer + 1

# Заполняем прочие расходы
    nomer = 7
    coll = 0
    categ = ""



    for i in other_db:

        if nomer > 24:
            nomer = 7
            coll = 3

        date = datetime.strptime((i[1]), '%Y-%m-%d %H:%M:%S')
        sheet.cell(row=nomer, column=9+coll).value = value = date.strftime('%d-%m-%y %H:%M')
        categ = emoji.replace_emoji(i[2], replace='')
        sheet.cell(row=nomer, column=11+coll).value = i[3]

        if i[4] != None:
            categ = categ + " " + f'"{emoji.replace_emoji(i[4], replace="")}"'
            sheet.cell(row=nomer, column=10 + coll).value = categ
        else:
            sheet.cell(row=nomer, column=10 + coll).value = categ
        nomer = nomer + 1


    filepath = f'Отчеты/{name}.xlsx'

    book.save(filepath)
    book.close()

    # excel_app = client.DispatchEx("Excel.Application")
    #
    # workbook = excel_app.Workbooks.Open(filepath)
    # worksheet = workbook.Worksheets(1)
    #
    # pdf_path = "output_file.pdf"
    #
    # worksheet.ExportAsFixedFormat(0, pdf_path)
    # workbook.Close(False)
    # excel_app.Quit()
    #
    return filepath
    # # file_dsc = pdf_path
    #

# # Обработка Эксель
@form_router.message(Command("exel"))
async def buy(message: types.Message, state: FSMContext):
    await state.set_state(orders_gen.namber)
    await message.answer("Введи номер рейса")

@form_router.message(orders_gen.namber)
async def get_info_m(message: types.Message, state: FSMContext) -> None:
    id = message.text

    ppyt = "C:\\Users\\Сергей\\PycharmProjects\\pythonProject2\\Отчеты\\"
    # ppyt = "Отчеты/"
    #
    data = await exel_creat(id)
    file_path = ppyt + data[7:]
    await state.clear()
    print("id fl {id}")
    print({id})
    # await message.bot.send_chat_action(
    #     chat_id=message.chat.id,
    #     action=ChatAction.UPLOAD_DOCUMENT,
    # )
    # # await message.answer(f"{data}")
    #
    # # WB_PATH3 = "C:\\Users\\Сергей\\PycharmProjects\\pythonProject2\\test.pdf"
    # WB_PATH3 = "test.pdf"
    #
    # file_dsc = "test.pdf"
    # excel = client.Dispatch("Excel.Application")
    # # # Read Excel File
    # sheets = excel.Workbooks.Open(ppyt)
    # work_sheets = sheets.Worksheets[0]
    # # # # Converting into PDF File
    # work_sheets.ExportAsFixedFormat(0, WB_PATH3)
    # sheets.Close(False)
    # excel.Quit()
    #
    # await message.reply_document(
    #     document=types.FSInputFile(
    #         path=file_dsc,
    #         filename="Отчёт.pdf"
    #     ),
    # )
    await message.answer("Создаю PDF...")

    WB_PATH3 = "C:\\PycharmProjects2\\Отчеты\\МазОт2024-01-18.xlsx"
    excel_file = "/Отчеты/МазОт2024-01-18.xlsx"  # Укажите путь к вашему файлу
    pdf_file = "C:\\PycharmProjects2\\Отчеты\\test777.pdf"

    try:
        subprocess.run(
            [r"C:\Program Files\LibreOffice\program\soffice.exe", "--headless", "--convert-to", "pdf", WB_PATH3, "--outdir", pdf_file],
            check=True
        )
        await message.reply_document(document=types.FSInputFile(pdf_file))
    except Exception as e:
        await message.answer(f"Ошибка при создании PDF: {e}")





